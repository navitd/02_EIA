import sys
from pathlib import Path
import os
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import openpyxl
import inspect
from openpyxl.cell.cell import MergedCell
# Add the parent directory to sys.path
sys.path.append(str(Path(__file__).resolve().parent.parent / 'EIAfunctions'))
from func_data_upload_OECD_salaries import data_upload_OECD_salaries
from func_plot_L import plot_matrix_columns
from func_clc_L import clc_L
from func_safe_divide import safe_divide, safe_divide_vector



def print_impacts_to_excel( year,
    direct_o: pd.DataFrame, indirect_o: pd.DataFrame, induced_o: pd.DataFrame, s2s_moc: pd.DataFrame,
    direct_h: pd.DataFrame, indirect_h: pd.DataFrame, induced_h: pd.DataFrame, s2s_mhc: pd.DataFrame,
    direct_g: pd.DataFrame, indirect_g: pd.DataFrame, induced_g: pd.DataFrame, s2s_mgc: pd.DataFrame,
    filename: str):

    def prepare_section(direct, indirect, induced, total, small_title):
        direct_sum = direct.T #direct.to_frame().T
        indirect_sum = indirect.T #indirect.to_frame().T
        induced_sum = induced.T #induced.to_frame().T
        total_sum = total.T#total.iloc[:-1, :-1].to_frame().T

    
        for df, label in zip(
            [direct_sum, indirect_sum, induced_sum, total_sum],
            ['Direct', 'Indirect', 'Induced', 'Total']):
            df.insert(0, small_title, [label])
        return pd.concat([direct_sum, indirect_sum, induced_sum, total_sum], ignore_index=True)

    output_df = prepare_section(direct_o, indirect_o, induced_o, s2s_moc, f"Output impact")
    income_df = prepare_section(direct_h, indirect_h, induced_h, s2s_mhc, f"Income impact")
    gdp_df = prepare_section(direct_g, indirect_g, induced_g, s2s_mgc, f"GDP impact")

    def write_section(ws, df, start_row, section_title):
        n_cols = df.shape[1]
        # Section title
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.value = section_title
        title_cell.font = Font(size=14, bold=True, color="000000")
        title_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column headers
        for col_num, column_title in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row + 1, column=col_num, value=column_title)
            if column_title in ["Output impact", "Income impact", "GDP impact"]:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
            else:
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
            cell.font = Font(bold=True)

        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:
                    cell.fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Light orange

        return start_row + 2 + len(df) + 1  # Next row start

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Dummy write to create sheet
        output_df.to_excel(writer, index=False, startrow=0)
        ws = writer.sheets['Sheet1']

        row = 1
        row = write_section(ws, output_df, row, f"{year} Output impact")
        row = write_section(ws, income_df, row, f"{year} Income impact")
        write_section(ws, gdp_df, row, f"{year} GDP impact")

    print(f"Excel file written to: {filename}")


def create_excel_file_with_title(year: str, filename: str = "output.xlsx") -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = year

    # Styles
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Merge title box over first 4 columns and 3 rows
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=4)
    cell = ws.cell(row=1, column=1)
    cell.value = f"EIA details - {year}"
    cell.fill = green
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = black_border

    wb.save(filename)

    return 1  # Next available column after title box

def append_styled_matrix_to_excel(df, matrix_name, year: str, start_col: int, filename: str = "output.xlsx", title_size=3) -> int:
    # Infer matrix name from variable name if not provided
    if matrix_name is None:
        frame = inspect.currentframe().f_back
        matrix_name = next((name for name, val in frame.f_locals.items() if val is df), "UnnamedMatrix")

    wb = openpyxl.load_workbook(filename)
    if year not in wb.sheetnames:
        raise ValueError(f"Sheet named '{year}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[year]

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Convert DataFrame to rows (including index and header)
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # includes index

    # Green title merged over up to 4 columns
    merge_end_col = min(start_col + title_size, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = matrix_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the matrix below the title
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)

    return sep_col + 1  # Return column to start the next matrix (skip separator too)

def append_styled_series_to_excel(series: pd.Series, series_name, year: str, start_col: int, filename: str = "output.xlsx") -> int:
    # Infer series name from variable name if not provided
    if series_name is None:
        frame = inspect.currentframe().f_back
        series_name = next((name for name, val in frame.f_locals.items() if val is series), "UnnamedSeries")

    wb = openpyxl.load_workbook(filename)
    if year not in wb.sheetnames:
        raise ValueError(f"Sheet named '{year}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[year]

    # Convert Series to DataFrame for uniformity
    df = series.to_frame(name=series.name if series.name else "Value")
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # Should be 2: index and value

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Green title merged over up to 2 columns
    merge_end_col = min(start_col + 1, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = series_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the series below the title
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)
    return sep_col + 1


def append_styled_multipliers_to_excel(
    year: str,
    direct_o: pd.Series, indirect_o: pd.Series, induced_o: pd.Series, total_o: pd.Series,
    direct_h: pd.Series, indirect_h: pd.Series, induced_h: pd.Series, total_h: pd.Series,
    direct_g: pd.Series, indirect_g: pd.Series, induced_g: pd.Series, total_g: pd.Series,
    filename: str,
    sheet_name: str
):
    def prepare_section(direct, indirect, induced, total, small_title):
        def build_row(series: pd.Series, label: str) -> pd.DataFrame:
            values = [float(v) for v in series.values]  # Ensure scalar floats
            return pd.DataFrame([[label] + values], columns=[small_title] + list(series.index))

        direct_df = build_row(direct, 'Direct')
        indirect_df = build_row(indirect, 'Indirect')
        induced_df = build_row(induced, 'Induced')
        total_df = build_row(total, 'Total')

        return pd.concat([direct_df, indirect_df, induced_df, total_df], ignore_index=True)

    output_df = prepare_section(direct_o, indirect_o, induced_o, total_o, f"Output impact")
    income_df = prepare_section(direct_h, indirect_h, induced_h, total_h, f"Income impact")
    gdp_df = prepare_section(direct_g, indirect_g, induced_g, total_g, f"GDP impact")

    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    def write_section(ws, df, start_row, section_title):
        n_cols = df.shape[1]
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.value = section_title
        title_cell.font = Font(size=14, bold=True, color="000000")
        title_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row + 1, column=col_num, value=column_title)
            if column_title in ["Output impact", "Income impact", "GDP impact"]:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 2):
            for col_idx, value in enumerate(row, start=1):
                if isinstance(value, (list, tuple, pd.Series, np.ndarray)):
                    value = float(value[0])  # Convert 1-element array/series to scalar
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:
                    cell.fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")

        return start_row + 2 + len(df) + 1

    row = 1
    row = write_section(ws, output_df, row, f"{year} Output impact")
    row = write_section(ws, income_df, row, f"{year} Income impact")
    write_section(ws, gdp_df, row, f"{year} GDP impact")

    wb.save(filename)
    print(f"Multipliers sheet added to: {filename}")



##################################################             old functions               ######################################################



def multipliers2prediction(s2s_mo, fdf_year2, column_name):
    predicted_output_year2_np  = np.round(s2s_mo.to_numpy() @ fdf_year2.values.reshape(-1, 1), 1)
    
    predicted_output_year2 = pd.DataFrame(predicted_output_year2_np, index=s2s_mo.index, columns=[column_name])
    
    return predicted_output_year2



def plot_market_multipliers(series_list, panel_titles, figure_title):
    
    fig, axes = plt.subplots(nrows=3, ncols=1, figsize=(6, 8), sharex=True)
    
    for ax, series, panel_title in zip(axes, series_list, panel_titles):
        ax.plot(series.index, series.values, marker='o', linestyle='-')
        ax.set_title(panel_title)
        ax.grid(True)
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='x', labelrotation=45)

    fig.tight_layout()
    fig.suptitle(figure_title, fontsize=14, y=0.98)
    plt.subplots_adjust(top=0.9)  # lower top to make room for suptitle
    plt.show()

def plot_real_vs_predicted(output_real, output_pred, 
                           income_real, income_pred, 
                           gdp_real, gdp_pred, 
                           year1, year2, title):
    fig, axes = plt.subplots(3, 1, figsize=(6,8), sharex=True)
    
    fig.suptitle(title, fontsize=16)

    # Panel 1: Output
    axes[0].plot(output_real.index, output_real, label='Real Output', color='purple', marker='o')
    axes[0].plot(output_pred.index, output_pred, label='Predicted Output', color='red', marker='o')
    axes[0].set_title(f'Output {year2} Based on {year1}')
    axes[0].set_xlabel('Sectors')
    axes[0].set_ylabel('Million USD')
    axes[0].legend()

    # Panel 2: Income
    axes[1].plot(income_real.index, income_real, label='Real Income', color='purple', marker='o')
    axes[1].plot(income_pred.index, income_pred, label='Predicted Income', color='red', marker='o')
    axes[1].set_title(f'Income {year2} Based on {year1}')
    axes[1].set_xlabel('Sectors')
    axes[1].set_ylabel('Million USD')
    axes[1].legend()

    # Panel 3: GDP
    axes[2].plot(gdp_real.index, gdp_real, label='Real GDP', color='purple', marker='o')
    axes[2].plot(gdp_pred.index, gdp_pred, label='Predicted GDP', color='red', marker='o')
    axes[2].set_title(f'GDP {year2} Based on {year1}')
    axes[2].set_xlabel('Sectors')
    axes[2].set_ylabel('Million USD')
    axes[2].legend()
    for ax in axes:
        ax.tick_params(axis='x', rotation=45)
        ax.grid(True)

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.show()

def plot_multipliers(OECD_sectors_ICT, direct_o, indirect_o, induced_o,
                     direct_h, indirect_h, induced_h,
                     direct_g, indirect_g, induced_g, title="Multipliers Plot"):
    
    fig, axes = plt.subplots(3, 1, figsize=(6, 8), sharex=True)
    fig.suptitle(title, fontsize=16)

    # Define bar width and positions
    bar_width = 0.25
    index = np.arange(len(OECD_sectors_ICT))

    # Panel 1: Output Multipliers
    axes[0].bar(index, direct_o, bar_width, label='Direct', color='green')
    axes[0].bar(index, indirect_o, bar_width, bottom=direct_o, label='Indirect', color='red')
    axes[0].bar(index, induced_o, bar_width, bottom=direct_o + indirect_o, label='Induced', color='blue')
    axes[0].set_title('Output Multipliers ICT sector')
    axes[0].set_ylabel('Multiplier Value')
    axes[0].set_xticks(index)
    axes[0].set_xticklabels(OECD_sectors_ICT, rotation=45)
    axes[0].legend()

    # Panel 2: Income Multipliers
    axes[1].bar(index, direct_h, bar_width, label='Direct', color='green')
    axes[1].bar(index, indirect_h, bar_width, bottom=direct_h, label='Indirect', color='red')
    axes[1].bar(index, induced_h, bar_width, bottom=direct_h + indirect_h, label='Induced', color='blue')
    axes[1].set_title('Income Multipliers ICT sector')
    axes[1].set_ylabel('Multiplier Value')
    axes[1].set_xticks(index)
    axes[1].set_xticklabels(OECD_sectors_ICT, rotation=45)
    axes[1].legend()

    # Panel 3: GDP Multipliers
    axes[2].bar(index, direct_g, bar_width, label='Direct', color='green')
    axes[2].bar(index, indirect_g, bar_width, bottom=direct_g, label='Indirect', color='red')
    axes[2].bar(index, induced_g, bar_width, bottom=direct_g + indirect_g, label='Induced', color='blue')
    axes[2].set_title('GDP Multipliers ICT sector')
    axes[2].set_xlabel('Sectors')
    axes[2].set_ylabel('Multiplier Value')
    axes[2].set_xticks(index)
    axes[2].set_xticklabels(OECD_sectors_ICT, rotation=45)
    axes[2].legend()

    # Adjust the layout for better visualization
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show()

 
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@                    main                  @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
start_time = time.time()
print("working directory of print2xls3.py is: ",os.getcwd())  # Print the current working directory

year = '2015'
year2 = '2015'
table_type = 'TTL' #or'DOM'   
OECD_path = "../Data/" # windows style: r".\\"
if table_type == 'DOM':
    output_filename = '/mnt/c/NavitComputer24/2024_NES/Economics/Textbook_EIA/OECD_salaries/EIA_matrices.xlsx'
elif table_type == 'TTL':
    output_filename = '/mnt/c/NavitComputer24/2024_NES/Economics/Textbook_EIA/OECD_salaries/EIA_TTL_matrices.xlsx'



# 1. Get IO=II, X, GDP, from OECD, compensation of employees, more GDP and II from OECDadditional as well as taxes, incomegross surplus etc.
##########################################################################################################################################   
currency_exchange_type = 'EXCH' #'EXCH' or 'PPP'
PPP_or_exch, OECD, simple_II_labels, OECDadditional, sector_description =  data_upload_OECD_salaries(year, currency_exchange_type, table_type)
print(f'PPP_or_exch {PPP_or_exch}')

additional_OECD_column_names = ['intermediate_consumption', 'mixed_income_gross', 'net_taxes_on_production',
                                'surplus_and_mixed_income_gross', 'output', 'salaries', 'employees_compensation', 'GDP' ]

# the following is calculated twice: in data_upload_OECD_salaries and here. I want to leave it here, but I also need it there - do I??
II = OECD.loc[simple_II_labels, simple_II_labels]
household_expenditure = OECD.loc[simple_II_labels, 'HFCE']
final_demand_columns = ['HFCE',	'NPISH',	'GGFC',	'GFCF',	'INVNT',	'CONS_NONRES', 'EXPO'] # 'IMPO', 'DPABR', 
other_final_demand = OECD.loc[simple_II_labels, final_demand_columns[1:]] #exluding HFCE - household expenditure
GDP         = OECD.loc['VALU', simple_II_labels]
output      = OECD.loc['OUTPUT', simple_II_labels]
#I don't need to worry about household_expenditure of GDP or output - they are both 0
# but output of GDP is given and should be marked independently

#single values in OECD:
#GDP_of_household_expenditure = OECD.loc['VALU', 'HFCE']
#GDP_of_total_column = OECD.loc['VALU', 'TOTAL'] # total is the output, should be equal to f_row_sums.sum(axis=0)
#GDP_of_final_demand = OECD.loc['VALU', final_demand_columns]        #this could be added to the rows from the right or to the columns form the bottom
#output_of_final_demand = OECD.loc['OUTPUT', final_demand_columns]   #probably will not be needed

# from looking at the numbers:
# total = output
# II_row_sums + f_row_sums = total = output
# household_expenditure is the column added to II to get the closed model
# when uploading data for the first time run:
# data_exploration_flags(II,household_expenditure,other_final_demand,output,output_of_final_demand,OECD_rough)

#final demand and value added are not the same at all

# 2. calculate L and Lc
###########################################################################################################################################
T = safe_divide(II, output)
Ldf, L_minus_I = clc_L(T)

IIc = II.copy()
IIc["HFCE"] = household_expenditure # added a column for closed model
IIc.loc['employees_compensation'] = OECDadditional['employees_compensation'] #If I wanted a column I would have written IIC['employees_compensation']
IIc.loc['employees_compensation', 'HFCE'] = 0 #T97_values.loc[T97_values['Transaction'] == 'Compensation of employees', 'OBS_VALUE_USD'].values[0]

outputc = output.copy()
outputc['HFCE'] = OECDadditional['employees_compensation'].sum()
Tc = safe_divide(IIc, outputc)
Lcdf, Lc_minus_I = clc_L(Tc)


# 3. calculate multipliers
###########################################################################################################################################

mo = Ldf.sum(axis=0) #dollar's worth of outcome per 1 dollar's worth of new final demand
moc_trancated = Lcdf.iloc[:-1].sum(axis=0) #dollar's worth of outcome per 1 dollar's worth of new final demand

# income multipliers mh
Ej_by_xj = Tc.iloc[-1,:-1] #hosehold income received per dollar's worth of sector output  

income_F_multipliers = Ldf.mul(Ej_by_xj, axis=0) #household income recieved per dollar's worth of secotr final demand
# Ej/xj*Ljk - Ljk is how much output was sold from j to k. and j is the sector that paid the salaries, so Ej/xj is used.
sum_income_F_multipliers = income_F_multipliers.sum(axis=0) 
# m(h)_k = sum_j(Ej/xj*Ljk) - sum over j of the detailed income_F_multipliers - sum over the rows
# an additional dolar of final demand in sector k generates m(h)_k dollars of new household income when all direct and
# indirect effects are converted into dollar estimates of income.
# income_F_multipliers is the details for each sector - how much income is generated by an additional dollar of final demand in sector k for each of the other sectors
# the above is only direct+indirect effects
# direct + indirect + induced effect - same calculation but with Lcdf

#income multipliers second time
Ej_by_xj = Tc.iloc[-1,:]
#it has a different size than above

# GDP multipliers
GDPc = OECD.loc['VALU', simple_II_labels + ['HFCE']]
GDPj_by_xj = safe_divide_vector(GDPc, outputc)

# summary of multipliers without typeI and typeII - 
# 12 multipliers output, income, GDP, X sector2sector, sector2market X simple model, closed model
# all of the closed model multipliers are trancated (the row and column of salaries and final demand are not included)
s2s_mo = Ldf                       # direct + indirect effect
s2s_moc = Lcdf                     # direct + indirect + iduced effect
s2s_mh = Ldf.mul(Ej_by_xj.iloc[ :-1 ], axis=0) 
s2s_mhc = Lcdf.mul(Ej_by_xj.rename(index={'HFCE': 'employees_compensation'}), axis=0)
s2s_mg =  Ldf.mul(GDPj_by_xj.iloc[ :-1 ], axis=0)    
s2s_mgc = Lcdf.mul(GDPj_by_xj.rename(index={'HFCE': 'employees_compensation'}), axis=0)

mo = s2s_mo.sum(axis=0)
moc = s2s_moc.sum(axis=0)
mh = s2s_mh.sum(axis=0)
mhc = s2s_mhc.sum(axis=0)
mg = s2s_mg.sum(axis=0)
mgc = s2s_mgc.sum(axis=0)


# impact analysis
###################################################
# multipliers: direct, indirect, induced separately
###################################################
n = T.shape[0]
# direct
direct_o = pd.DataFrame(np.eye(n), index=s2s_mo.index, columns=s2s_mo.columns)
direct_h = pd.DataFrame(np.zeros((n, n)), index=Ej_by_xj.iloc[:-1].index, columns=Ej_by_xj.iloc[:-1].index)
np.fill_diagonal(direct_h.values, Ej_by_xj.values)
direct_g = pd.DataFrame(np.zeros((n, n)), index=GDPj_by_xj.iloc[:-1].index, columns=GDPj_by_xj.iloc[:-1].index)
np.fill_diagonal(direct_g.values, GDPj_by_xj.values)
#indirect
indirect_o = s2s_mo - direct_o
#Ej_by_xj*L_minus_I = s2s_mh-Ej_by_xj
indirect_h  = s2s_mh - direct_h
#GDPj_by_xj*L_minus_I = s2s_mg-GDPj_by_xj
indirect_g  = s2s_mg - direct_g
#induced
induced_o = s2s_moc.iloc[:-1,:-1] - s2s_mo
induced_h = s2s_mhc.iloc[:-1,:-1] - s2s_mh
induced_g = s2s_mgc.iloc[:-1,:-1] - s2s_mg


start_col = create_excel_file_with_title(year, filename=output_filename)
start_col = append_styled_matrix_to_excel(IIc, 'IIc', year, start_col, filename=output_filename) 
start_col = append_styled_series_to_excel(outputc, 'outputc', year, start_col, filename=output_filename )
start_col = append_styled_series_to_excel(GDP, 'OECD GDP', year, start_col, filename=output_filename )
start_col = append_styled_matrix_to_excel(T, 'T', year, start_col, filename=output_filename )
start_col = append_styled_matrix_to_excel(Tc, 'Tc', year, start_col, filename=output_filename)
start_col = append_styled_matrix_to_excel(Ldf, 'L', year, start_col, filename=output_filename)
start_col = append_styled_matrix_to_excel(Lcdf, 'Lc', year, start_col, filename=output_filename)
# output multipliers divided into sectors
start_col = append_styled_matrix_to_excel(direct_o, 'sector to sector output multiliers, direct', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_mo, 'sector to sector output multiliers, direct+indirect', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_moc, 'sector to sector output multiliers, direct+indirct+induced', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(direct_h, 'sector to sector E multiliers, direct', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_mh, 'sector to sector E multiliers, direct+indirect', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_mhc, 'sector to sector E multiliers, direct+indirct+induced', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(direct_g, 'sector to sector GDP multiliers, direct', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_mg, 'sector to sector GDP multiliers, direct+indirect', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_mgc, 'sector to sector GDP multiliers, direct+indirct+induced', year, start_col, filename=output_filename, title_size=6)






append_styled_multipliers_to_excel( year,
                       direct_o.sum(axis=0), indirect_o.sum(axis=0), induced_o.sum(axis=0), s2s_moc.iloc[:-1, :-1].sum(axis=0),
                       direct_h.sum(axis=0), indirect_h.sum(axis=0), induced_h.sum(axis=0), s2s_mhc.iloc[:-1, :-1].sum(axis=0),
                       direct_g.sum(axis=0), indirect_g.sum(axis=0), induced_g.sum(axis=0), s2s_mgc.iloc[:-1, :-1].sum(axis=0),
                       filename=output_filename,
                        sheet_name = f'{year} Multipliers' )




# predict output, income and GDP
#################################
#year2 = '2015'

_, OECD_year2, _, OECDadditional_year2, _ =  data_upload_OECD_salaries(year, currency_exchange_type, table_type)
income_year2 = OECDadditional_year2['employees_compensation']
GDP_year2 = OECD_year2.loc['VALU', simple_II_labels]

fdf_year2 = OECD_year2.loc[simple_II_labels, final_demand_columns].sum(axis=1)
#there is what causes closed model to be in accuarete:
#fcdf_year2 = OECD_year2.loc[simple_II_labels,final_demand_columns[1:]].sum(axis=1)
#I should take HFCE inside fcdf_year2. 
fcdf_year2 = OECD_year2.loc[simple_II_labels,final_demand_columns].sum(axis=1)
fcdf_year2.loc['employees_compensation'] = 0

predicted_output_year2 = multipliers2prediction(s2s_mo, fdf_year2, 'Predicted_Output')
predicted_outputc_year2 = multipliers2prediction(s2s_moc, fcdf_year2, 'Predicted_Output')
predicted_income_year2 = multipliers2prediction(s2s_mh, fdf_year2, 'Predicted_Income')  
predicted_incomec_year2 = multipliers2prediction(s2s_mhc, fcdf_year2, 'Predicted_Income') 
predicted_GDP_year2 = multipliers2prediction(s2s_mg, fdf_year2, 'Predicted_GDP') 
predicted_GDPc_year2 = multipliers2prediction(s2s_mgc, fcdf_year2, 'Predicted_GDP') 
output_year2      = OECD_year2.loc['OUTPUT', simple_II_labels]


def multipliers_by_f(M, fcdf_year2, title):
    fcdf_year2 = fcdf_year2.values.reshape(-1, 1) if isinstance(fcdf_year2, pd.Series) else fcdf_year2
    result = M.values @ fcdf_year2
    result_df = pd.DataFrame(result, index=M.index, columns=[title])
    return result_df

temp = multipliers_by_f(s2s_moc.iloc[:-1,:-1], fcdf_year2[:-1], 'Total output impact')


append_styled_multipliers_to_excel( year,
                       multipliers_by_f(direct_o, fcdf_year2[:-1], 'Direct output impact'), 
                       multipliers_by_f(indirect_o, fcdf_year2[:-1], 'Indirect output impact'),
                       multipliers_by_f(induced_o, fcdf_year2[:-1], 'Induced output impact'),  
                       multipliers_by_f(s2s_moc.iloc[:-1,:-1], fcdf_year2[:-1], 'Total output impact'),
                       multipliers_by_f(direct_h, fcdf_year2[:-1], 'Direct income impact'), 
                       multipliers_by_f(indirect_h, fcdf_year2[:-1], 'Indirect income impact'),
                       multipliers_by_f(induced_h, fcdf_year2[:-1], 'Induced income impact'),  
                       multipliers_by_f(s2s_mhc.iloc[:-1,:-1], fcdf_year2[:-1], 'Total income impact'),
                       multipliers_by_f(direct_g, fcdf_year2[:-1], 'Direct GDP impact'), 
                       multipliers_by_f(indirect_g, fcdf_year2[:-1], 'Indirect GDP impact'),
                       multipliers_by_f(induced_g, fcdf_year2[:-1], 'Induced GDP impact'),  
                       multipliers_by_f(s2s_mgc.iloc[:-1,:-1], fcdf_year2[:-1], 'Total GDP impact'),  
                       filename=output_filename,
                       sheet_name = f"{year} Impacts")




# going back to the first worksheet and adding the impacts
def s2s_impacts_multipliers_by_f(M, fcdf_year2):
    fcdf_year2 = fcdf_year2.values.reshape(1, -1) if isinstance(fcdf_year2, pd.Series) else fcdf_year2.reshape(1, -1)
    result = M.values * fcdf_year2  # Element-wise multiplication, broadcasted across rows
    result_df = pd.DataFrame(result, index=M.index, columns=M.columns)
    return result_df

# output impacts divided into sectors
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(direct_o, fcdf_year2.iloc[:-1]), 'sector to sector output impact, direct', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(s2s_mo, fcdf_year2.iloc[:-1]), 'sector to sector output impact, direct+indirect', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(s2s_moc, fcdf_year2), 'sector to sector output impact, direct+indirct+induced', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(direct_h, fcdf_year2.iloc[:-1]), 'sector to sector E impact, direct', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(s2s_mh, fcdf_year2.iloc[:-1]), 'sector to sector E impact, direct+indirect', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(s2s_mhc, fcdf_year2), 'sector to sector E impact, direct+indirct+induced', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(direct_g, fcdf_year2.iloc[:-1]), 'sector to sector GDP impact, direct', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(s2s_mg, fcdf_year2.iloc[:-1]), 'sector to sector GDP impact, direct+indirect', year, start_col, filename=output_filename, title_size=6)
start_col = append_styled_matrix_to_excel(s2s_impacts_multipliers_by_f(s2s_mgc, fcdf_year2), 'sector to sector GDP impact, direct+indirct+induced', year, start_col, filename=output_filename, title_size=6)


# calculate type I and typeII multipliers
print()



##############################              plotting          #############################
'''
plot_real_vs_predicted(output_year2, predicted_output_year2,
                       income_year2, predicted_income_year2,
                       GDP_year2, predicted_GDP_year2,  
                       year, year2,'Simple Model')



plot_real_vs_predicted(output_year2, predicted_outputc_year2.iloc[:-1],
                       income_year2, predicted_incomec_year2.iloc[:-1],
                       GDP_year2, predicted_GDPc_year2.iloc[:-1],  
                       year, year2,'Closed Model')

'''
                       
                       
'''                    
# see how Lcdf compares to L
plot_heatmap(Lc_minus_I.iloc[:-1,:-1] - L_minus_I, f"Lc-L, Lc trancated, {year}")
'''



#multipliers plotting
'''
plot_market_multipliers([mo, mh, mg], ['New Dollar\'s Output per New Dollar\'s Final Demand',
                                       'New Dollar\'s Income per New Dollar\'s Final Demand',
                                       'New Dollar\'s GDP per New Dollar\'s Final Demand'], figure_title=f"{year}, Simple Model: Direct + Indirect")
plot_market_multipliers([moc, mhc, mgc], ['New Dollar\'s Output per New Dollar\'s Final Demand',
                                          'New Dollar\'s Income per New Dollar\'s Final Demand',
                                          'New Dollar\'s GDP per New Dollar\'s Final Demand'], figure_title=f"{year}, Closed Model: Direct + Indirect + Induced")



plot_market_multipliers([s2s_mo.loc[:,sector], s2s_mh.loc[:,sector], s2s_mg.loc[:,sector]], ['New Dollar\'s Output per New Dollar\'s Final Demand',
                                       'New Dollar\'s Income per New Dollar\'s Final Demand',
                                       'New Dollar\'s GDP per New Dollar\'s Final Demand'], 
                                       figure_title=f"{year}, Simple Model: Direct + Indirect, {sector_description}")
plot_market_multipliers([s2s_moc.loc[:,sector], s2s_mhc.loc[:,sector], s2s_mgc.loc[:,sector]], ['New Dollar\'s Output per New Dollar\'s Final Demand',
                                       'New Dollar\'s Income per New Dollar\'s Final Demand',
                                       'New Dollar\'s GDP per New Dollar\'s Final Demand'], 
                                       figure_title=f"{year}, Closed Model: Direct + Indirect + Induced, {sector_description}")

'''

# bar graphs of direct, indirect and induced
'''
ICT_sectors = ['ICT - Manufacturing', 'ICT - Wholesaling', 'ICT - Software and computer services', 'ICT - Communications services',
               'ICT - Software and computer services',	'ICT - Software and computer services']
OECD_sectors_ICT = ['C26',	'G',	'J58T60',	'J61',	'J62_63',	'M']
ICT_sectors_dict = {'ICT - Manufacturing': 'C26',
                    'ICT - Wholesaling': 'G',
                    'ICT - Software and computer services': ['J58T60', 'J62_63', 'M'],  
                    'ICT - Communications services': 'J61'}
# Build sector code to name mapping
code_to_name = {}
for name, codes in ICT_sectors_dict.items():
    if isinstance(codes, list):
        for code in codes:
            code_to_name[code] = name
    else:
        code_to_name[codes] = name

# assume I want to see the ICT sectors as selling sectors. how much they will sell
plot_multipliers(OECD_sectors_ICT, direct_o.loc[OECD_sectors_ICT,:].sum(axis=1), indirect_o.loc[OECD_sectors_ICT,:].sum(axis=1), 
                 induced_o.loc[OECD_sectors_ICT,:].sum(axis=1),
                 direct_h.loc[OECD_sectors_ICT,:].sum(axis=1), indirect_h.loc[OECD_sectors_ICT,:].sum(axis=1),
                 induced_h.loc[OECD_sectors_ICT,:].sum(axis=1),
                 direct_g.loc[OECD_sectors_ICT,:].sum(axis=1), indirect_g.loc[OECD_sectors_ICT,:].sum(axis=1),
                 induced_g.loc[OECD_sectors_ICT,:].sum(axis=1), 
                  title="Multipliers")
'''






# the following is an old plotting. only ICT and only T
'''
# Modul 3: plotting
#ICT sectors information
ICT_sectors = ['ICT - Manufacturing', 'ICT - Wholesaling', 'ICT - Software and computer services', 'ICT - Communications services',
               'ICT - Software and computer services',	'ICT - Software and computer services']
# These correspond to the numbers 17	26	33	34	35	38
OECD_sectors_ICT = ['C26',	'G',	'J58T60',	'J61',	'J62_63',	'M']
ICT_sectors_dict = {'ICT - Manufacturing': 'C26',
                    'ICT - Wholesaling': 'G',
                    'ICT - Software and computer services': ['J58T60', 'J62_63', 'M'],  
                    'ICT - Communications services': 'J61'}
# Build sector code to name mapping
code_to_name = {}
for name, codes in ICT_sectors_dict.items():
    if isinstance(codes, list):
        for code in codes:
            code_to_name[code] = name
    else:
        code_to_name[codes] = name



##Simple output multipliers: L
#plot_matrix_columns(
#    matrix=Ldf,
#    sectors=OECD_sectors_ICT,
#    sector_code_to_name=code_to_name,
#    title=f'Leontief Matrix Column Profiles - output direct+indirect impact, year {year}'
#)



plot_matrix_columns(
    matrix=L_minus_I,
    sectors=OECD_sectors_ICT,
    sector_code_to_name=code_to_name,
    title=f'output indirect impact, year {year}'
)'
'''