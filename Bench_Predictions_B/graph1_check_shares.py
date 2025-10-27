# standartization of data collection
# what I have:
# 1995-2010: OECD II + E extrap
# 2011-2020: OECD II + OECD E
# 2021-2040: Lc extrap, E extrap
# extrap = extrapolated, mainly by gdp data from world bank. there's ARIMA in gdp and linear extrapolation in japan gdp

#In this file I will upload everything, and make necessary alternations to dataframe so that future years data and data years data is the same

# https://www.oecd.org/en/data/datasets/input-output-tables.html



#I know I have GDP extrapolated from world bank - but - there is 10% difference between world bank GDP and OECD GDP
#so I used fixed year GDP 2020 divided by output, as per EIA method, and will multiply by extrapolated output of each future year




import sys
from pathlib import Path
import os
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import matplotlib.gridspec as gridspec
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.cell.cell import MergedCell
import inspect

# Add the parent directory to sys.path
sys.path.append(str(Path(__file__).resolve().parent.parent / 'EIAfunctions'))
from func_data_upload_OECD_without_E import data_upload_OECD_without_E
from func_plot_L import plot_matrix_columns
from func_clc_L import clc_L
from func_safe_divide import safe_divide, safe_divide_vector
from func_multipliers_by_f import multipliers_by_f
from func_plot_real_vs_predicted import plot_real_vs_predicted


##########################################         functions from Benchmarking/Employment.py       ##########################################
#the plotting functions from Employment are not needed here, they are to be moved to 08 file

# collect_v and collect_m - this is moving from 1year to many years - collecting in a dataframe an aggregation of the differnt years
def collect_v(v, country, year, cols_list, dfv):
    dftemp = pd.DataFrame()
    dftemp = v.reset_index()
    dftemp.columns = cols_list 
    dftemp['country'] = country
    dftemp['year'] = year
    dftemp = dftemp[["country", "year"] + cols_list]
    dfv = pd.concat([dfv, dftemp], ignore_index=True)
    return dfv

def collect_m(m, country, year, m_value_name, dfm):
    dftemp = pd.DataFrame()
    dftemp = m.reset_index().melt(id_vars=m.index.name or 'index', 
                                    var_name='buying_sector', 
                                    value_name=m_value_name)

    # Rename 'index' to 'selling_sector' if needed
    dftemp.rename(columns={m.index.name or 'index': 'selling_sector'}, inplace=True)
    # Add metadata
    dftemp['country'] = country
    dftemp['year'] = year
    # Reorder columns
    dftemp = dftemp[['country', 'year',  'selling_sector', 'buying_sector', m_value_name]]
    # Append to the master DataFrame
    dfm = pd.concat([dfm, dftemp], ignore_index=True)
    return dfm

def slice_v_from_bigdf(bigdf,country,year):
    # assuming bigdf was prepared by collect_v: columns are country, year, sector, value_column
    v = bigdf[(bigdf.country==country) & (bigdf.year==int(year))].copy()
    #remove country and year from E and add 0 at the end [employees_compensation, HFCE]=0
    v.drop(columns=['country','year'], inplace=True)
    v.set_index('sector', inplace=True)
    return v


####################################################         functions that plot       ######################################################
def plot_E_line_graph(JPNE, col_name, title):
    years = sorted(JPNE['year'].unique())
    num_years = len(years)

    # Generate colors from red to purple using the 'rainbow' colormap
    colors = cm.rainbow(np.linspace(0, 1, num_years))

    plt.figure(figsize=(14, 6))

    for i, year in enumerate(years):
        data = JPNE[JPNE['year'] == year]
        plt.plot(data['sector'], data['Employment'], label=str(year), color=colors[i])

    plt.xlabel('Sector')
    plt.ylabel(col_name)
    plt.title(title)
    plt.xticks(rotation=90)
    plt.legend(title='Year')
    plt.tight_layout()
    plt.show()

#Fig 1: CAGR data manipulation
def clc_cagr(dfoutput, first_year, last_year, value_column):
    # pivoting is a great idea there's no groupby. there's just taking first_year and last_year, then pivoting to ahve each row isolate an equation, then we manipulate numbers in each row
    df_filtered = dfoutput[dfoutput['year'].isin([first_year, last_year])]
    pivot_df = df_filtered.pivot_table(
        index=['country', 'sector'],
        columns='year',
        values=value_column
    ).reset_index()
    #np.where is actually an if statement np.where(condition, value_if_true, value_if_false)
    pivot_df['CAGR'] = np.where(
        (pivot_df[first_year] != 0) & (pivot_df[first_year].notna()) & (pivot_df[last_year].notna()),
        (pivot_df[last_year] / pivot_df[first_year]) ** (1 / (int(last_year) - int(first_year))) - 1,
        np.nan
    )
    #plotting - take all the ICT sectors and average CAGR, then plot.
    ICT_cagr = (
        pivot_df[pivot_df['sector'].isin(ICTsectors)]   # Filter for ICT sectors
        .groupby('country')['CAGR']                     # group by country
        .mean()                                         # calculate mean CAGR for each country                   
        .sort_values(ascending=False)
    )
    return ICT_cagr

# fig 1: plot CAGR
def plot_cagr(ICT_cagr, title):
        # Rename for plotting
    ICT_cagr.index = [country_map[c] for c in ICT_cagr.index]

    # Plot
    fig, ax = plt.subplots(figsize=(10, 6))
    x = np.arange(len(ICT_cagr))
    bars = ax.bar(x, ICT_cagr.values, color=[
        'red' if country_map.get(code, code) == 'Canada' else 'blue' for code in ICT_cagr.index
    ])

    # Add labels
    for i, bar in enumerate(bars):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, height + 0.002,
                f'{height*100:.1f}%', ha='center', va='bottom', fontsize=9)

    # Formatting
    ax.set_xticks(x)
    ax.set_xticklabels([country_map.get(code, code) for code in ICT_cagr.index], rotation=45, ha='right')
    ax.set_ylabel('Average CAGR [%]')
    ax.set_title(title)

    plt.tight_layout()
    plt.show()


# Fig 2: output share data manipulation
def get_share(dfoutput, first_year, last_year, ICTsectors, value_column):
    #the output row from OECD is the output needed according to Tanveer - it includes import and HFCE etc. I checked.
    df_filtered2 = dfoutput.pivot_table(
    index=['country', 'sector'],
    columns='year',
    values=value_column
    ).reset_index()

    yearly_output = dfoutput.groupby(['country', 'year'])[value_column].sum().reset_index()
    yearly_output = yearly_output.rename(columns={value_column: f'total yearly {value_column}'})

    for year in range(int(first_year), int(last_year) + 1):
        # merge yearly_output with df_filtered2 to get total output for each country and year
        df_filtered2 = df_filtered2.merge(
            yearly_output[yearly_output['year'] == year][['country', f'total yearly {value_column}']],
            on='country',
            how='left'
        )
        df_filtered2 = df_filtered2.rename(columns={f'total yearly {value_column}': f'total yearly {value_column} {year}'})

        df_filtered2[f'{value_column} share {year}'] = df_filtered2[year] / df_filtered2[f'total yearly {value_column} {year}']

    # Define the years
    years = list(range(int(first_year), int(last_year) + 1))

    # Build the list of output share columns
    share_cols = [f'{value_column} share {year}' for year in years]

    # df_filtered2 has output per sector per year per country and also the ratio of that sector output to total yearly output
    shares = df_filtered2[['country', 'sector'] + share_cols].copy()

    # Calculate the average share across the selected years
    shares[f'average_{value_column}_share'] = shares[share_cols].mean(axis=1)

    ICT_shares = shares[shares['sector'].isin(ICTsectors)][['country', 'sector', f'average_{value_column}_share']].copy()

    return shares, ICT_shares

# plot fig 2: output share
def plot_share(ICT_shares, title,value_column):
    # Group by country and sort
    country_avg = ICT_shares.groupby('country')[f'average_{value_column}_share'].mean().sort_values(ascending=False)

    # Define colors
    colors = ['red' if country == 'CAN' else 'blue' for country in country_avg.index]

    # Plot
    plt.figure(figsize=(10, 6))
    bars = plt.bar(country_avg.index, country_avg.values, color=colors)

    # Adjust y-axis limit for label space
    max_height = country_avg.max()
    plt.ylim(0, max_height * 1.15)

    # Add percentage labels
    for bar in bars:
        height = bar.get_height()
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            height + max_height * 0.015,
            f'{height * 100:.1f}%',
            ha='center',
            va='bottom',
            fontsize=9
        )

    plt.ylabel(f'Average ICT {value_column} Share (%)')
    plt.title(title)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.show()

# fig2B: plot stacked output share
#def plot_stacked_output_shares(output_shares, ICT_factors, title):
def plot_stacked_shares(shares, ICT_factors, title, value_column):

    # Invert dictionary ICT_factors
    sector_to_category = {}
    for category, sectors in ICT_factors.items():
        if isinstance(sectors, list):
            for sector in sectors:
                sector_to_category[sector] = category
        else:
            sector_to_category[sectors] = category

    # Filter for ICT sectors
    ICTsectors = list(sector_to_category.keys())
    ICT_shares = shares.loc[shares['sector'].isin(ICTsectors), ['country', 'sector', f'average_{value_column}_share']].copy()

    # Map to ICT category
    ICT_shares['ICT_category'] = ICT_shares['sector'].map(sector_to_category)

    # Group by country and ICT_category, sum average_output_share
    grouped = ICT_shares.groupby(['country', 'ICT_category'])[f'average_{value_column}_share'].sum().unstack(fill_value=0)
                                                                                            # country and ICT_category are the index. unstack will create columns for each ICT_category
                                                                                            # fill_value=0 will fill NaN with 0
    # Reorder columns for consistent stacking: bottom to top
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling', 'ICT - Software and computer services', 'ICT - Communications services']

    # Sum across ICT categories to get total ICT share per country, then sort descending
    grouped['total'] = grouped.sum(axis=1)
    grouped = grouped.sort_values('total', ascending=False).drop(columns='total')

    # Now `countries` is updated to match the new order
    countries = grouped.index.tolist()

    # Continue with plotting as before
    colors = ['#4CAF50', '#CD7F32', '#FFC107', '#2196F3'] 
    bottom = np.zeros(len(countries))
    plt.figure(figsize=(10, 6))

    for idx, category in enumerate(desired_order):
        values = grouped[category].values
        #bars = plt.bar(countries, values, bottom=bottom, color=colors[idx], label=category)
        bars = plt.bar(countries, values * 100, bottom=bottom * 100, color=colors[idx], label=category)
        bottom += values
        
    # Add % labels on top
    for i, total in enumerate(bottom):
        plt.text(i, total * 100 + 0.2, f"{total * 100:.1f}%", ha='center', va='bottom', fontsize=9)

    #plt.ylabel('Average ICT Output Share')
    plt.ylabel(f'Average ICT {value_column} Share (%)')
    plt.title(title)
    plt.xticks(rotation=45, ha='right')
    plt.legend(title="ICT Category", bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.tight_layout()
    plt.show()

    
#fig 2C: plot ICT GDP stacked share, comparison of first and last year

def plot_share_compare_frist_last_year(shares, first_year, last_year, value_column, title):

    # Invert dictionary ICT_factors
    sector_to_category = {}
    for category, sectors in ICT_factors.items():
        if isinstance(sectors, list):
            for sector in sectors:
                sector_to_category[sector] = category
        else:
            sector_to_category[sectors] = category

    # Filter for ICT sectors
    ICTsectors = list(sector_to_category.keys())
    ICT_shares_first_year = shares.loc[shares['sector'].isin(ICTsectors), ['country', 'sector', f'GDP share {first_year}']].copy()
    ICT_shares_last_year = shares.loc[shares['sector'].isin(ICTsectors), ['country', 'sector', f'GDP share {last_year}']].copy()
    # Map to ICT category
    ICT_shares_first_year['ICT_category'] = ICT_shares_first_year['sector'].map(sector_to_category)
    ICT_shares_last_year['ICT_category'] = ICT_shares_last_year['sector'].map(sector_to_category)

    # Group by country and ICT_category, sum average_output_share
    grouped_first_year = ICT_shares_first_year.groupby(['country', 'ICT_category'])[f'GDP share {first_year}'].sum().unstack(fill_value=0)
    grouped_last_year = ICT_shares_last_year.groupby(['country', 'ICT_category'])[f'GDP share {last_year}'].sum().unstack(fill_value=0)

                                                                                            # country and ICT_category are the index. unstack will create columns for each ICT_category
                                                                                            # fill_value=0 will fill NaN with 0
    # Reorder columns for consistent stacking: bottom to top
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling', 'ICT - Software and computer services', 'ICT - Communications services']

    # Sum across ICT categories to get total ICT share per country, then sort descending
    grouped_first_year['total'] = grouped_first_year.sum(axis=1)
    grouped_last_year['total'] = grouped_last_year.sum(axis=1)
    grouped_first_year = grouped_first_year.sort_values('total', ascending=False).drop(columns='total')
    grouped_last_year = grouped_last_year.sort_values('total', ascending=False).drop(columns='total')

    # Now `countries` is updated to match the new order
    countries = grouped_last_year.index.tolist()

    # Setup
    countries = grouped_first_year.index.tolist()
    n_countries = len(countries)
    x = np.arange(n_countries)  # X positions for the bars
    bar_width = 0.35

    # Colors
    base_colors = ['#4CAF50', '#2196F3', '#FFC107', '#9C27B0']  # vivid for last_year

    # Create faded colors for first_year
    def fade_color(hex_color, blend=0.4):
        rgb = np.array(mcolors.to_rgb(hex_color))
        white = np.ones_like(rgb)
        faded_rgb = rgb * (1 - blend) + white * blend
        return faded_rgb

    faded_colors = [fade_color(c) for c in base_colors]

    # Plotting
    fig, ax = plt.subplots(figsize=(12, 6))
    bottom_first = np.zeros(n_countries)
    bottom_last = np.zeros(n_countries)

    for idx, category in enumerate(desired_order):
        values_first = grouped_first_year[category].values
        values_last = grouped_last_year[category].values

        ax.bar(x - bar_width / 2, values_first * 100, bottom=bottom_first * 100,
            color=faded_colors[idx], width=bar_width, label=f"{category} ({first_year})" if idx == 0 else "", alpha=0.8)

        ax.bar(x + bar_width / 2, values_last * 100, bottom=bottom_last * 100,
            color=base_colors[idx], width=bar_width, label=f"{category} ({last_year})" if idx == 0 else "")

        bottom_first += values_first
        bottom_last += values_last

    # Add % labels above bars for total (optional)
    for i in range(n_countries):
        ax.text(x[i] - bar_width / 2, bottom_first[i] * 100 + 1, f"{bottom_first[i] * 100:.1f}%", ha='center', fontsize=8)
        ax.text(x[i] + bar_width / 2, bottom_last[i] * 100 + 1, f"{bottom_last[i] * 100:.1f}%", ha='center', fontsize=8)

    # Final plot setup
    ax.set_ylabel(f'Average ICT {value_column} Share (%)')
    ax.set_title(title)
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45, ha='right')
    # Add more space above the highest bar
    max_height = max(np.max(bottom_first), np.max(bottom_last)) * 100
    ax.set_ylim(top=max_height * 1.1)  # 10% extra space above tallest bar

    # Custom legend (merged by category)
    custom_legend = [Patch(color=faded_colors[i], label=f"{cat} ({first_year})") for i, cat in enumerate(desired_order)]
    custom_legend += [Patch(color=base_colors[i], label=f"{cat} ({last_year})") for i, cat in enumerate(desired_order)]
    ax.legend(handles=custom_legend, bbox_to_anchor=(1.05, 1), loc='upper left', title="ICT Category")

    plt.tight_layout()
    plt.show()
# fig 3
def get_one_year_value(df, year, forward_or_backward, sector_list, value_column):
    if forward_or_backward == 'backward':
        col = 'buying sector'
    else:
        col = 'selling sector'
    name = value_column.split()[0]

    #first slicing 
    one_year_impact = df[
        (df['year'] == year) & 
        (df[col].isin(sector_list))       
    ][['country', 'year', 'selling sector', 'buying sector', value_column, f'national {name}']]
    # division
    one_year_impact[value_column] = ( one_year_impact[value_column] / one_year_impact[f'national {name}'] )
    one_year_impact.drop(columns=[f'national {name}'], inplace=True)

    one_year_impact_grouped = one_year_impact.groupby(['country', 'year'], as_index=False)[value_column].sum()
    one_year_impact_grouped = one_year_impact_grouped.sort_values(by=f'{name} impact total', ascending=False)
    return one_year_impact_grouped


def plot_GDPimpact_side_by_side(
    first_year_backwards, last_year_backwards,
    first_year_forwards, last_year_forwards,
    value_column, graph_title):
    width = 0.35
    fig, axes = plt.subplots(1, 2, figsize=(16, 6), sharey=True)

    # --- Panel 1: Backward Linkage ---
    ax = axes[0]
    last_year_backwards = last_year_backwards.sort_values(by=value_column, ascending=False)
    countries = last_year_backwards['country'].values

    first_year_backwards = first_year_backwards.set_index('country').loc[countries].reset_index()
    last_year_backwards = last_year_backwards.reset_index(drop=True)

    x = np.arange(len(countries))

    # Define colors
    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_backwards[value_column], width,
                   label=f"{first_year_backwards['year'].iloc[0]}", color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_backwards[value_column], width,
                   label=f"{last_year_backwards['year'].iloc[0]}", color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Backward Linkage')
    ax.set_xlabel('Country')
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)
    ax.legend()

    # --- Panel 2: Forward Linkage ---
    ax = axes[1]
    last_year_forwards = last_year_forwards.sort_values(by=value_column, ascending=False)
    countries = last_year_forwards['country'].values

    first_year_forwards = first_year_forwards.set_index('country').loc[countries].reset_index()
    last_year_forwards = last_year_forwards.reset_index(drop=True)

    x = np.arange(len(countries))

    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_forwards[value_column], width,
                   label=f"{first_year_forwards['year'].iloc[0]}", color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_forwards[value_column], width,
                   label=f"{last_year_forwards['year'].iloc[0]}", color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Forward Linkage')
    ax.set_xlabel('Country')
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)
    ax.legend()

    fig.supylabel(value_column)
    fig.suptitle(f'{graph_title}, {value_column}: Comparison of {first_year_backwards["year"].iloc[0]} and {last_year_backwards["year"].iloc[0]}')
    plt.tight_layout()
    plt.show()


def plot_GDPimpact_top_bottom(
    first_year_backwards, last_year_backwards,
    first_year_forwards, last_year_forwards,
    value_column, graph_title):

    width = 0.35
    fig, axes = plt.subplots(2, 1, figsize=(12, 10), sharex=False)

    # --- Panel 1: Backward Linkage ---
    ax = axes[0]
    last_year_backwards = last_year_backwards.sort_values(by=value_column, ascending=False)
    countries = last_year_backwards['country'].values

    first_year_backwards = first_year_backwards.set_index('country').loc[countries].reset_index()
    last_year_backwards = last_year_backwards.reset_index(drop=True)

    x = np.arange(len(countries))
    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_backwards[value_column], width, color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_backwards[value_column], width, color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Backward Linkage')
    ax.set_ylabel(value_column)
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)

    # Custom legend (light blue = first year, dark blue = last year)
    custom_legend = [
        Patch(facecolor='skyblue', label=str(first_year_backwards['year'].iloc[0])),
        Patch(facecolor='navy', label=str(last_year_backwards['year'].iloc[0]))
    ]
    ax.legend(handles=custom_legend)

    # --- Panel 2: Forward Linkage ---
    ax = axes[1]
    last_year_forwards = last_year_forwards.sort_values(by=value_column, ascending=False)
    countries = last_year_forwards['country'].values

    first_year_forwards = first_year_forwards.set_index('country').loc[countries].reset_index()
    last_year_forwards = last_year_forwards.reset_index(drop=True)

    x = np.arange(len(countries))
    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_forwards[value_column], width, color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_forwards[value_column], width, color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Forward Linkage')
    ax.set_ylabel(value_column)
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)
    ax.legend(handles=custom_legend)

    fig.suptitle(f'{graph_title}, {value_column}: Comparison of {first_year_backwards["year"].iloc[0]} and {last_year_backwards["year"].iloc[0]}')
    plt.tight_layout()
    plt.show()



# fig. 3 for ICT categories
def plot_GDPimpact_wrapper(dfGDPimpact, first_year, last_year, sector_list, value_column, sector_label):
    #this function uses the plot function above to plot 4 different categories of the ICT sector
    # Extract data for backward and forward linkages
    first_year_backward = get_one_year_value(dfGDPimpact, first_year, 'backward', sector_list, value_column)
    last_year_backward = get_one_year_value(dfGDPimpact, last_year, 'backward', sector_list, value_column)
    first_year_forward = get_one_year_value(dfGDPimpact, first_year, 'forward', sector_list, value_column)
    last_year_forward = get_one_year_value(dfGDPimpact, last_year, 'forward', sector_list, value_column)

    # Plot using existing plotting function
    plot_GDPimpact_top_bottom(first_year_backward, last_year_backward,
                              first_year_forward, last_year_forward,
                             'GDP impact total', sector_label)
    # Optionally set a custom plot title
    #plt.suptitle(f'{value_column} for {sector_label}', fontsize=14)
    #plt.tight_layout()
    #plt.show()


# fig. 4 GDP impact backward forward stacked
def plot_stacked_ict_impact(backward_df, forward_df, year, value_col, title):
    
    # Filter for the given year and reindex
    back = backward_df[backward_df['year'] == year][['country', value_col]].set_index('country')
    fwd = forward_df[forward_df['year'] == year][['country', value_col]].set_index('country')

    # Align and fill missing values
    back, fwd = back.align(fwd, join='outer', fill_value=0)

    # Sort by total impact
    total = back[value_col] + fwd[value_col]
    sorted_countries = total.sort_values(ascending=False).index

    # Sort the data
    back_sorted = back.loc[sorted_countries]
    fwd_sorted = fwd.loc[sorted_countries]

    # Create the bar plot
    fig, ax = plt.subplots(figsize=(12, 6))
    for i, country in enumerate(sorted_countries):
        back_val = back_sorted.loc[country, value_col]
        fwd_val = fwd_sorted.loc[country, value_col]
        total_val = back_val + fwd_val

        # Choose colors
        if country == 'CAN':
            back_color = 'forestgreen'
            fwd_color = 'lightgreen'
        else:
            back_color = 'indigo'
            fwd_color = 'orchid'

        # Plot bars
        ax.bar(i, back_val, color=back_color)
        ax.bar(i, fwd_val, bottom=back_val, color=fwd_color)

        # Add percentage label
        ax.text(i, total_val + 0.01, f'{total_val * 100:.1f}%', ha='center', va='bottom', fontsize=9)

    # Final plot settings
    ax.set_title(f"{title} in {year}", fontsize=14)
    ax.set_ylabel('GDP Impact')
    ax.set_xlabel('Country')
    ax.set_xticks(range(len(sorted_countries)))
    ax.set_xticklabels(sorted_countries, rotation=45)

    # Legend
    from matplotlib.patches import Patch
    legend_handles = [
        Patch(color='indigo', label='Backward (others)'),
        Patch(color='orchid', label='Forward (others)'),
        Patch(color='forestgreen', label='Backward (CAN)'),
        Patch(color='lightgreen', label='Forward (CAN)')
    ]
    ax.legend(handles=legend_handles)

    ax.grid(axis='y', linestyle='--', alpha=0.5)
    plt.tight_layout()
    plt.show()




##################################################           print to excel            ############################################

def create_excel_file_with_title(ws_title: str, filename) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = ws_title

    # Styles
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Merge title box over first 4 columns and 3 rows
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=4)
    cell = ws.cell(row=1, column=1)
    cell.value = f"EIA details - {ws_title}"
    cell.fill = green
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = black_border

    wb.save(filename)

    return 1  # Next available column after title box


def append_styled_matrix_to_excel(df, matrix_name, worksheet_name, start_col: int, filename, title_size=3) -> int:
    # Infer matrix name from variable name if not provided
    if matrix_name is None:
        frame = inspect.currentframe().f_back
        matrix_name = next((name for name, val in frame.f_locals.items() if val is df), "UnnamedMatrix")
    
    wb = load_workbook(filename)
    if worksheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet named '{worksheet_name}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[worksheet_name]

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Convert DataFrame to rows (including index and header)
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # includes index

    # Green title merged over up to 4 columns
    merge_end_col = min(start_col + title_size, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = matrix_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the matrix below the title
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)

    return sep_col + 1  # Return column to start the next matrix (skip separator too)



def append_styled_series_to_excel(series: pd.Series, series_name, year: str, start_col: int, filename: str = "output.xlsx") -> int:
    # Infer series name from variable name if not provided
    if series_name is None:
        frame = inspect.currentframe().f_back
        series_name = next((name for name, val in frame.f_locals.items() if val is series), "UnnamedSeries")

    wb = load_workbook(filename)
    if year not in wb.sheetnames:
        raise ValueError(f"Sheet named '{year}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[year]

    # Convert Series to DataFrame for uniformity
    df = series.to_frame(name=series.name if series.name else "Value")
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # Should be 2: index and value

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Green title merged over up to 2 columns
    merge_end_col = min(start_col + 1, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = series_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the series below the title
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)
    return sep_col + 1

##################################################        functions that calculate        ######################################################

def multipliers2prediction(s2s_mo, fdf_year2, column_name):
    predicted_output_year2_np  = np.round(s2s_mo.to_numpy() @ fdf_year2.values.reshape(-1, 1), 1)
    
    predicted_output_year2 = pd.DataFrame(predicted_output_year2_np, index=s2s_mo.index, columns=[column_name])
    
    return predicted_output_year2


def scale_df_by_series(direct_o: pd.DataFrame, fcdf: pd.Series) -> pd.DataFrame:
    
    return direct_o[fcdf.index].mul(fcdf, axis=1)


def pivot_matrix_to_3_columns(m: pd.DataFrame, value: str) -> pd.DataFrame:
    return m.reset_index().melt(id_vars=m.index.name or 'index',
                                var_name='buying sector',
                                value_name=value).rename(columns={m.index.name or 'index': 'selling sector'})


def get_impacts(dfimpact, mdirect, mindirect, minduced, ms2s, value_vec, value_vec_name, value_col, country,year):
    impact_cols = [value_col+' impact direct', value_col+' impact indirect', value_col+' impact induced', value_col+' impact total']
    dftemp2 = None
    for data, value in zip( [mdirect, mindirect, minduced, ms2s], impact_cols ):
        m = scale_df_by_series(data, fcdf[:-1])       #this is the multiplication
        dftemp1 = pivot_matrix_to_3_columns(m, value) #this is the matrix in 3 columns
        if dftemp2 is None:
            dftemp2 = dftemp1  # First iteration: just assign
        else:
            dftemp2 = pd.merge(
                dftemp2,
                dftemp1,
                on=["selling sector", "buying sector"],
                how="outer"
            )
    # dftemp2 contains 4 GDP impacts 
    dftemp2['country'] = country
    dftemp2['year'] = year
    dftemp2[value_vec_name] = value_vec.sum()
    cols = ['country', 'year', 'buying sector', 'selling sector'] + impact_cols + [value_vec_name]
    dftemp2 = dftemp2[cols]
    dfimpact = pd.concat([dfimpact, dftemp2], ignore_index=True)
    return dfimpact




def plot_stacked_shares_with_printing(shares, ICT_factors, title, value_column,worksheet_name, start_col, filename):

    # Invert dictionary ICT_factors
    sector_to_category = {}
    for category, sectors in ICT_factors.items():
        if isinstance(sectors, list):
            for sector in sectors:
                sector_to_category[sector] = category
        else:
            sector_to_category[sectors] = category

    # Filter for ICT sectors
    ICTsectors = list(sector_to_category.keys())
    ICT_shares = shares.loc[shares['sector'].isin(ICTsectors), ['country', 'sector', f'average_{value_column}_share']].copy()

    # Map to ICT category
    ICT_shares['ICT_category'] = ICT_shares['sector'].map(sector_to_category)

    # Group by country and ICT_category, sum average_output_share
    grouped = ICT_shares.groupby(['country', 'ICT_category'])[f'average_{value_column}_share'].sum().unstack(fill_value=0)
                                                                                            # country and ICT_category are the index. unstack will create columns for each ICT_category
                                                                                            # fill_value=0 will fill NaN with 0
    # Reorder columns for consistent stacking: bottom to top
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling', 'ICT - Software and computer services', 'ICT - Communications services']

    # Sum across ICT categories to get total ICT share per country, then sort descending
    grouped['total'] = grouped.sum(axis=1)
    grouped = grouped.sort_values('total', ascending=False).drop(columns='total')

    start_col = append_styled_matrix_to_excel(grouped, 'grouped', worksheet_name, start_col, filename=xlsx_filename)

    # Now `countries` is updated to match the new order
    countries = grouped.index.tolist()

    # Continue with plotting as before
    colors = ['#4CAF50', '#CD7F32', '#FFC107', '#2196F3'] 
    bottom = np.zeros(len(countries))
    plt.figure(figsize=(10, 6))

    for idx, category in enumerate(desired_order):
        values = grouped[category].values
        #bars = plt.bar(countries, values, bottom=bottom, color=colors[idx], label=category)
        bars = plt.bar(countries, values * 100, bottom=bottom * 100, color=colors[idx], label=category)
        bottom += values
        
    # Add % labels on top
    for i, total in enumerate(bottom):
        plt.text(i, total * 100 + 0.2, f"{total * 100:.1f}%", ha='center', va='bottom', fontsize=9)

    #plt.ylabel('Average ICT Output Share')
    plt.ylabel(f'Average ICT {value_column} Share (%)')
    plt.title(title)
    plt.xticks(rotation=45, ha='right')
    plt.legend(title="ICT Category", bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.tight_layout()
    plt.show()

    return start_col


#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@                    main                  @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# upload gdp
dfgdp_worldbank = pd.read_csv("Bench_predictions_B/A04_gdp_ARIMAgdp_currentUSD04.csv")
worldbank_gdp_col_name = "gdp total world bank"
dfgdp_worldbank.rename(columns={"Unnamed: 0": "year","gdp total": worldbank_gdp_col_name}, inplace=True)
dfgdp_worldbank.iloc[:, 1:] = dfgdp_worldbank.iloc[:, 1:] * 10**(-6)
dfgdp_worldbank['year'] = dfgdp_worldbank['year'].astype(int)
dfgdp_worldbank = dfgdp_worldbank.set_index('year')


# upload E
dfE = pd.read_csv("Bench_predictions_B/A05_Esectors_from_Etot05.csv")
dfE.rename(columns={"E": "Employment"}, inplace=True)
# dfE already has data until 2040 - base extrapolation already done.

# for base extrapolation
dfTc = pd.read_csv("Bench_predictions_B/B071_Tc_base_1years.csv")
dfHFCE = pd.read_csv("Bench_predictions_B/B072_dfHFCE_data_and_extrap.csv")
df8    = pd.read_csv("Bench_predictions_B/B072_df8_data_and_extrap.csv")
df9    = pd.read_csv("Bench_predictions_B/B072_df9_data_and_extrap.csv")
dfGDP  = pd.read_csv("Bench_predictions_B/B072_dfGDP_data_and_extrap.csv")
dfoutput = pd.read_csv("Bench_predictions_B/B072_dfoutput_data_and_extrap.csv")
dfGDPj_by_xj = pd.read_csv("Bench_predictions_B/B072_dfGDPj_by_xj_data_and_extrap.csv")

#dfHFCE_tot       = pd.read_csv("Bench_predictions_B/B06_dfHFCE_tot.csv")
#df8_tot          = pd.read_csv("Bench_predictions_B/B06_df8_tot.csv")
#df9_tot          = pd.read_csv("Bench_predictions_B/B06_df9_tot.csv")
#dfGDP_tot        = pd.read_csv("Bench_predictions_B/B06_dfGDP_tot.csv")
#dfoutput_tot     = pd.read_csv("Bench_predictions_B/B06_dfoutput_tot.csv") # needed to get future years
#dfGDPj_by_xj_tot = pd.read_csv("Bench_predictions_B/B06_dfGDPj_by_xj_tot.csv")

########################################                           parameters                       ##################################################
start_time = time.time()
print("working directory of B10_graphs1.py is: ",os.getcwd())  # Print the current working directory

table_type = 'TTL' #or'DOM'   
OECD_path = "../Data/" # windows style: r".\\"
if table_type == 'DOM':
    output_filename = '/mnt/c/NavitComputer24/2024_NES/Economics/Textbook_EIA/OECD_salaries/EIA_matrices.xlsx'
    final_demand_columns = ['HFCE',	'NPISH', 'GGFC',	'GFCF',	'INVNT', 'CONS_NONRES', 'EXPO']
elif table_type == 'TTL':
    output_filename = '/mnt/c/NavitComputer24/2024_NES/Economics/Textbook_EIA/OECD_salaries/EIA_TTL_matrices.xlsx'
    final_demand_columns = ['HFCE', 'NPISH', 'GGFC', 'GFCF', 'INVNT', 'DPABR', 'CONS_NONRES', 'EXPO', 'IMPO']

# data years
first_year = 2014 
last_year = 2024  
year_range = [int(year) for year in range(int(first_year), int(last_year) + 1)]
# 
# possible
# make a function out of B07_base_sectros so that I can run it with parameters (n_years_for_base for example)

# future years
#max_future_year = dfoutput_tot.year.unique().max()
#year_range_future = [int(year) for year in range(int(last_year+1), max_future_year+1)]

#report_title = f'ICT sectors, {last_year}'
ICT_factors = {'ICT - Manufacturing': 'C26',
                'ICT - Wholesaling': 'G',
                'ICT - Software and computer services': ['J58T60', 'J62_63', 'M'],  
                'ICT - Communications services': 'J61'}
ICTsectors = ['C26', 'G', 'J58T60', 'J62_63', 'M', 'J61']

country_names = ['Canada', 'The United States', 'Great Britain', 'France', 'Germany', 'Italiy', 'Japan']
countries = ['CAN', 'USA', 'GBR', 'FRA', 'DEU', 'ITA', 'JPN'] # 'CHN' is not available in OECD, but it is in OECDadditional
country_map = dict(zip(countries, country_names))

currency_exchange_type = 'EXCH' #'EXCH' or 'PPP'

simple_II_labels = ['A01_02', 'A03', 'B05_06', 'B07_08', 'B09', 'C10T12', 'C13T15', 'C16', 'C17_18', 'C19', 'C20', 'C21', 'C22', 'C23', 'C24', 
                 'C25', 'C26', 'C27', 'C28', 'C29', 'C30', 'C31T33', 'D', 'E', 'F', 'G', 'H49', 'H50', 'H51', 'H52', 'H53', 'I', 'J58T60', 'J61',
                  'J62_63', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']

first_year_4graph = 2014
last_year_4graph = 2024
dfoutput_4graph = dfoutput[(dfoutput.year >= first_year_4graph) & (dfoutput.year <= last_year_4graph) ].copy()


##########################################             Benchmark  plots            ######################################################


print(f'Fig 1: ICT Sector Revenue Compound Annual Growth Rate (CAGR) ({first_year}-{last_year})')


# fig 2: average ICT sector share in total output 2010-2020
# data manipulation for figure 2: output share of ICT sectors
output_shares, ICT_output_shares = get_share(dfoutput_4graph, first_year_4graph, last_year_4graph, ICTsectors,'output')

#print to an excel file
xlsx_filename = "Bench_predictions_B/B10 output shares graph1 data.xlsx"
worksheet_name = f"output shares {first_year_4graph}-{last_year_4graph}"
start_col = create_excel_file_with_title(worksheet_name, xlsx_filename )
start_col = append_styled_matrix_to_excel(output_shares, 'output_shares', worksheet_name, start_col, filename=xlsx_filename)
start_col = append_styled_matrix_to_excel(ICT_output_shares, 'ICT_output_shares', worksheet_name, start_col, filename=xlsx_filename)

#country_avg = ICT_output_shares.groupby('country')[f'average_output_share'].mean().sort_values(ascending=False)
#start_col = append_styled_matrix_to_excel(country_avg, 'country average', worksheet_name, start_col+1, filename=xlsx_filename)
for year in year_range:
    for country in countries: 
        start_col = append_styled_matrix_to_excel(dfoutput[(dfoutput.country==country) & (dfoutput.year==year)], 'output', worksheet_name, start_col, filename=xlsx_filename )


plot_share(ICT_output_shares, f'Average ICT Output Share by Country, {first_year_4graph}-{last_year_4graph}','output')
#the above is average of average - average over the 6 ICT sectorsa as well as over the years

# fig2B: stacked output share
#this is the average of each category (factor) - stacked. 
start_col = plot_stacked_shares_with_printing(output_shares, ICT_factors,f'Stacked Average ICT Output Share by Country, {first_year_4graph}-{last_year_4graph}','output',worksheet_name, start_col, xlsx_filename)



# graphs 1 and 2 for GDP
if 0:
    # fig 1: output CAGR 
    ICT_GDP_cagr = clc_cagr(dfGDP, first_year, last_year,'GDP') 
    # fig1: plot output CAGR
    plot_cagr(ICT_GDP_cagr, f'Average GDP CAGR for ICT sectors ({first_year}–{last_year})')

    # fig 2: average ICT sector share in GDP 2011-2020
    GDP_shares, ICT_GDP_shares = get_share(dfGDP, first_year, last_year, ICTsectors,'GDP')
    plot_share(ICT_GDP_shares, f'Average ICT GDP Share by Country, {first_year}-{last_year}','GDP')

    # fig2B: stacked output share
    #this is the average of each category (factor) - stacked. 
    plot_stacked_shares(GDP_shares, ICT_factors,f'Stacked Average ICT GDP Share by Country, {first_year}-{last_year}','GDP')

#GDP share stacked, not average but comparison between 2011 and 2020
if 0:
    GDP_shares, ICT_GDP_shares = get_share(dfGDP, first_year, last_year, ICTsectors,'GDP')
    plot_share_compare_frist_last_year(GDP_shares, first_year, last_year, 'GDP', f'ICT GDP {first_year} and {last_year} Share by Country')


print('graphs 1 and 2 are done')

