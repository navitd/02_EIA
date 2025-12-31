# standartization of data collection
# what I have:
# 1995-2010: OECD II + E extrap
# 2011-2020: OECD II + OECD E
# 2021-2040: Lc extrap, E extrap
# extrap = extrapolated, mainly by gdp data from world bank. there's ARIMA in gdp and linear extrapolation in japan gdp
# https://www.oecd.org/en/data/datasets/input-output-tables.html




import sys
from pathlib import Path
import os
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import matplotlib.gridspec as gridspec
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.cell.cell import MergedCell
import inspect
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter



# Add the parent directory to sys.path
sys.path.append(str(Path(__file__).resolve().parent.parent / 'EIAfunctions'))
from func_data_upload_OECD_without_E import data_upload_OECD_without_E
from func_plot_L import plot_matrix_columns
from func_clc_L import clc_L
from func_safe_divide import safe_divide, safe_divide_vector
from func_plot_real_vs_predicted import plot_real_vs_predicted
       

def get_vector_impacts(country, year, multipliers, f8, f9, value_name, induced_flag='f8'):
        #correct:
        #f8.rename(index={"employees_compensation":"HFCE"},inplace=True) 
        #f9.rename(index={"employees_compensation":"HFCE"},inplace=True) 
        #multipliers are only for these country and year:
        country = multipliers['country']
        year = multipliers['year']
        if value_name =="output":
            names = ["direct_o", "indirect_o", "induced_o"]
        elif value_name =="GDP":
            names = ["direct_g", "indirect_g", "induced_g"]
        elif value_name =="Employment":
            names = ["direct_h", "indirect_h", "induced_h"]

        direct = multipliers[["selling sector", "buying sector", names[0]]].copy().pivot_table(index=['selling sector'], columns='buying sector',
        values=names[0])
        indirect = multipliers[["selling sector", "buying sector", names[1]]].copy().pivot_table(index=["selling sector"], columns="buying sector", 
                                                                                                        values=names[1]) 
        induced = multipliers[["selling sector", "buying sector", names[2]]].copy().pivot_table(index=["selling sector"], columns="buying sector", 
                                                                                                        values=names[2])
        # note that in the following I multiply L by f8 and not f9. this is so that they sum to output.
        f8.rename(index={"employees_compensation":"HFCE"},inplace=True) 
        f9.rename(index={"employees_compensation":"HFCE"},inplace=True) 
        
        if induced_flag=='f8': #to check that output = sum of impacts
            direct_impact = multipliers2prediction(direct, f8.drop("HFCE"), value_name)
            indirect_impact = multipliers2prediction(indirect, f8.drop("HFCE"), value_name)
            induced_impact = multipliers2prediction(induced, f8.drop("HFCE"), value_name)
        elif induced_flag=='f9': # real caclulation
            direct_impact = multipliers2prediction(direct, f9.drop("HFCE"), value_name)
            indirect_impact = multipliers2prediction(indirect, f9.drop("HFCE"), value_name)
            induced_impact = multipliers2prediction(induced, f8.drop("HFCE"), value_name)
        # induced is trancated - not n+1 but n rows/columns
        return  country, year, direct_impact, indirect_impact, induced_impact


#######################################        functions that plot B version from B10        #####################################   

#Fig 1: CAGR data manipulation
def clc_cagr(dfdata, first_year, last_year, value_column):
    # pivoting is a great idea there's no groupby. there's just taking first_year and last_year, then pivoting to ahve each row isolate an equation, then we manipulate numbers in each row
    df_filtered = dfdata[dfdata['year'].isin([first_year, last_year])]
    pivot_df = df_filtered.pivot_table(
        index=['country', 'sector'],
        columns='year',
        values=value_column
    ).reset_index()
    #np.where is actually an if statement np.where(condition, value_if_true, value_if_false)
    pivot_df['CAGR'] = np.where(
        (pivot_df[first_year] != 0) & (pivot_df[first_year].notna()) & (pivot_df[last_year].notna()),
        (pivot_df[last_year] / pivot_df[first_year]) ** (1 / (int(last_year) - int(first_year))) - 1,
        np.nan
    )
    #plotting - take all the ICT sectors and average CAGR, then plot.
    ICT_cagr = (
        pivot_df[pivot_df['sector'].isin(ICTsectors)]   # Filter for ICT sectors
        .groupby('country')['CAGR']                     # group by country
        .mean()                                         # calculate mean CAGR for each country                   
        .sort_values(ascending=False)
    )
    return ICT_cagr

# fig 1: plot CAGR
def plot_cagr(ICT_cagr, title):
    # the following is for changing 'CAN' to Canada, and I can only do it once, so if I plot and embed it throws an error
    #ICT_cagr.index = [country_map[c] for c in ICT_cagr.index]

    fig, ax = plt.subplots(figsize=(10, 6))
    x = np.arange(len(ICT_cagr))
    bars = ax.bar(x, ICT_cagr.values, color=[
        'red' if country_map.get(code, code) == 'Canada' else 'blue' for code in ICT_cagr.index
    ])

    for i, bar in enumerate(bars):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, height + 0.002,
                f'{height*100:.1f}%', ha='center', va='bottom', fontsize=9)

    ax.set_xticks(x)
    ax.set_xticklabels([country_map.get(code, code) for code in ICT_cagr.index], rotation=45, ha='right')
    ax.set_ylabel('Average CAGR [%]')
    ax.set_title(title)
    fig.tight_layout()

    return fig


# Fig 2: output share data manipulation
def get_share(dfoutput, value_column, first_year, last_year, ICTsectors):
    #the output row from OECD is the output needed according to Tanveer - it includes import and HFCE etc. I checked.
    df_filtered2 = dfoutput.pivot_table(
    index=['country', 'sector'],
    columns='year',
    values=value_column
    ).reset_index()
    # I calculate the total and the rate here
    yearly_output = dfoutput.groupby(['country', 'year'])[value_column].sum().reset_index()
    yearly_output = yearly_output.rename(columns={value_column: f'total yearly {value_column}'})

    for year in range(int(first_year), int(last_year) + 1):
        
        # merge yearly_output with df_filtered2 to get total output for each country and year
        df_filtered2 = df_filtered2.merge(
            yearly_output[yearly_output['year'] == year][['country', f'total yearly {value_column}']],
            on='country',
            how='left'
        )
        df_filtered2 = df_filtered2.rename(columns={f'total yearly {value_column}': f'total yearly {value_column} {year}'})

        df_filtered2[f'{value_column} share {year}'] = df_filtered2[year] / df_filtered2[f'total yearly {value_column} {year}']

    # Define the years
    years = list(range(int(first_year), int(last_year) + 1))

    # Build the list of output share columns
    share_cols = [f'{value_column} share {year}' for year in years]

    # df_filtered2 has output per sector per year per country and also the ratio of that sector output to total yearly output
    shares = df_filtered2[['country', 'sector'] + share_cols].copy()

    # Calculate the average share across the selected years
    shares[f'average_{value_column}_share'] = shares[share_cols].mean(axis=1)

    ICT_shares = shares[shares['sector'].isin(ICTsectors)][['country', 'sector', f'average_{value_column}_share']].copy()

    return shares, ICT_shares

# fig2B: plot stacked output share
#old version of plot_stacked_shares is in B10_graphs1
def get_share_by_category(shares, value_column, ICTcategories, desired_order):
    sector_to_category = {}
    for category, sectors in ICTcategories.items():
        if isinstance(sectors, list):
            for sector in sectors:
                sector_to_category[sector] = category
        else:
            sector_to_category[sectors] = category

    ICTsectors = list(sector_to_category.keys())
    ICT_shares = shares.loc[shares['sector'].isin(ICTsectors),
                            ['country', 'sector', f'average_{value_column}_share']].copy()
    ICT_shares['ICT_category'] = ICT_shares['sector'].map(sector_to_category)

    grouped = (ICT_shares
               .groupby(['country', 'ICT_category'])[f'average_{value_column}_share']
               .sum()
               .unstack(fill_value=0))

    
    grouped['total'] = grouped.sum(axis=1)
    grouped = grouped.sort_values('total', ascending=False).drop(columns='total')
    

    grouped = grouped[desired_order].copy()
    return grouped


def plot_stacked_shares(shares, value_column, ICTcategories, desired_order, title, highlighted):
    
    countries = shares.index.tolist()
    category_colors = {}
    for category, sectors in ICTcategories.items():
        if isinstance(sectors, list):
            color_hex = None
            for s in sectors:
                if s in highlighted:
                    color_hex = "#" + highlighted[s].lstrip("#")
                    break
            category_colors[category] = color_hex or "#CCCCCC"
        else:
            s = sectors
            color_hex = "#" + highlighted.get(s, "CCCCCC").lstrip("#")
            category_colors[category] = color_hex

    fig, ax = plt.subplots(figsize=(10, 6))
    bottom = np.zeros(len(countries))

    for category in desired_order:
        if category in shares.columns:
            values = shares[category].values
            color = category_colors.get(category, "#CCCCCC")
            ax.bar(countries, values * 100, bottom=bottom * 100,
                   color=color, label=category)
            bottom += values

    for i, total in enumerate(bottom):
        ax.text(i, total * 100 + 0.2, f"{total * 100:.1f}%",
                ha='center', va='bottom', fontsize=9)

    ax.set_ylabel(f'Average ICT {value_column} Share (%)')
    ax.set_title(title)
    ax.set_xticks(range(len(countries)))
    ax.set_xticklabels(countries, rotation=45, ha='right')
    ax.legend(title="ICT Category", bbox_to_anchor=(1.05, 1), loc='upper left')
    fig.tight_layout()

    return fig

#fig 2C: plot ICT GDP stacked share, comparison of first and last year
def plot_share_compare_first_last_year(shares, value_column, first_year, last_year, title):

    # Invert dictionary ICTcategories
    sector_to_category = {}
    for category, sectors in ICTcategories.items():
        if isinstance(sectors, list):
            for sector in sectors:
                sector_to_category[sector] = category
        else:
            sector_to_category[sectors] = category

    # Filter for ICT sectors
    ICTsectors = list(sector_to_category.keys())
    ICT_shares_first_year = shares.loc[shares['sector'].isin(ICTsectors), ['country', 'sector', f'{varname} share {first_year}']].copy()
    ICT_shares_last_year = shares.loc[shares['sector'].isin(ICTsectors), ['country', 'sector', f'{varname} share {last_year}']].copy()
    # Map to ICT category
    ICT_shares_first_year['ICT_category'] = ICT_shares_first_year['sector'].map(sector_to_category)
    ICT_shares_last_year['ICT_category'] = ICT_shares_last_year['sector'].map(sector_to_category)

    # Group by country and ICT_category, sum average_output_share
    grouped_first_year = ICT_shares_first_year.groupby(['country', 'ICT_category'])[f'{varname} share {first_year}'].sum().unstack(fill_value=0)
    grouped_last_year = ICT_shares_last_year.groupby(['country', 'ICT_category'])[f'{varname} share {last_year}'].sum().unstack(fill_value=0)

                                                                                            # country and ICT_category are the index. unstack will create columns for each ICT_category
                                                                                            # fill_value=0 will fill NaN with 0
    # Reorder columns for consistent stacking: bottom to top
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling', 'ICT - Software and computer services', 'ICT - Communications services']

    # Sum across ICT categories to get total ICT share per country, then sort descending
    grouped_first_year['total'] = grouped_first_year.sum(axis=1)
    grouped_last_year['total'] = grouped_last_year.sum(axis=1)
    grouped_first_year = grouped_first_year.sort_values('total', ascending=False).drop(columns='total')
    grouped_last_year = grouped_last_year.sort_values('total', ascending=False).drop(columns='total')

    # Now `countries` is updated to match the new order
    countries = grouped_last_year.index.tolist()

    # Setup
    countries = grouped_first_year.index.tolist()
    n_countries = len(countries)
    x = np.arange(n_countries)  # X positions for the bars
    bar_width = 0.35

    # Colors
    base_colors = ['#4CAF50', '#2196F3', '#FFC107', '#9C27B0']  # vivid for last_year

    # Create faded colors for first_year
    def fade_color(hex_color, blend=0.4):
        rgb = np.array(mcolors.to_rgb(hex_color))
        white = np.ones_like(rgb)
        faded_rgb = rgb * (1 - blend) + white * blend
        return faded_rgb

    faded_colors = [fade_color(c) for c in base_colors]

    # Plotting
    fig, ax = plt.subplots(figsize=(12, 6))
    bottom_first = np.zeros(n_countries)
    bottom_last = np.zeros(n_countries)

    for idx, category in enumerate(desired_order):
        values_first = grouped_first_year[category].values
        values_last = grouped_last_year[category].values

        ax.bar(x - bar_width / 2, values_first * 100, bottom=bottom_first * 100,
            color=faded_colors[idx], width=bar_width, label=f"{category} ({first_year})" if idx == 0 else "", alpha=0.8)

        ax.bar(x + bar_width / 2, values_last * 100, bottom=bottom_last * 100,
            color=base_colors[idx], width=bar_width, label=f"{category} ({last_year})" if idx == 0 else "")

        bottom_first += values_first
        bottom_last += values_last

    # Add % labels above bars for total (optional)
    for i in range(n_countries):
        ax.text(x[i] - bar_width / 2, bottom_first[i] * 100 + 1, f"{bottom_first[i] * 100:.1f}%", ha='center', fontsize=8)
        ax.text(x[i] + bar_width / 2, bottom_last[i] * 100 + 1, f"{bottom_last[i] * 100:.1f}%", ha='center', fontsize=8)

    # Final plot setup
    ax.set_ylabel(f'Average ICT {value_column} Share (%)')
    ax.set_title(title)
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45, ha='right')
    # Add more space above the highest bar
    max_height = max(np.max(bottom_first), np.max(bottom_last)) * 100
    ax.set_ylim(top=max_height * 1.1)  # 10% extra space above tallest bar

    # Custom legend (merged by category)
    custom_legend = [Patch(color=faded_colors[i], label=f"{cat} ({first_year})") for i, cat in enumerate(desired_order)]
    custom_legend += [Patch(color=base_colors[i], label=f"{cat} ({last_year})") for i, cat in enumerate(desired_order)]
    ax.legend(handles=custom_legend, bbox_to_anchor=(1.05, 1), loc='upper left', title="ICT Category")

    plt.tight_layout()
    plt.show()



##################################################           print to excel fancy version B         ############################################

def create_excel_file_with_title(ws_title: str, filename) -> int:

    if os.path.exists(filename):
        wb = load_workbook(filename)
        if ws_title in wb.sheetnames:
            ws = wb[ws_title]
        else:
            ws = wb.create_sheet(title=ws_title)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = ws_title
    
    # Styles
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Merge title box over first 4 columns and 3 rows
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=4)
    cell = ws.cell(row=1, column=1)
    cell.value = f"EIA details - {ws_title}"
    cell.fill = green
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = black_border

    wb.save(filename)

    return 1  # Next available column after title box

def append_styled_matrix_to_excel(df, matrix_name, worksheet_name, start_col, filename, highlighted_sectors, title_size ) -> int:
    
    # Infer matrix name from variable name if not provided
    if matrix_name is None:
        frame = inspect.currentframe().f_back
        matrix_name = next((name for name, val in frame.f_locals.items() if val is df), "UnnamedMatrix")
    
    wb = load_workbook(filename)
    if worksheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet named '{worksheet_name}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[worksheet_name]

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    fill_styles = {}
    if highlighted_sectors:
        for sector, color in highlighted_sectors.items():
            fill_styles[sector] = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Convert DataFrame to rows (including index and header)
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # includes index

    # Green title merged over up to 4 columns
    merge_end_col = min(start_col + title_size, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = matrix_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the matrix below the title
    for r_idx, row in enumerate(rows, start=5):
        fill_color = None
        if r_idx > 5 and highlighted_sectors:  # skip header
            try:
                sector_col_idx = df.columns.get_loc("sector") + 1  # +1 for index col
                sector_value = row[sector_col_idx]
                if sector_value in fill_styles:
                    fill_color = fill_styles[sector_value]
            except Exception:
                pass

        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)

            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            elif fill_color:
                cell.fill = fill_color

            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)
    return sep_col + 1

# the following is a special printing of the category matrix. it is different because categoris are in columns, not rows. it prints well.
def append_styled_matrix_by_category_to_excel(df, filename, worksheet_name, matrix_name, start_col, ICTcategories, highlighted, title_size):
    
    # Infer matrix name if not provided
    if matrix_name is None:
        frame = inspect.currentframe().f_back
        matrix_name = next((name for name, val in frame.f_locals.items() if val is df), "UnnamedMatrix")

    wb = load_workbook(filename)
    if worksheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet named '{worksheet_name}' does not exist.")
    ws = wb[worksheet_name]

    # === Styles ===
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # === Derive category colors from ICTcategories and highlighted ===
    category_colors = {}
    for category, sectors in ICTcategories.items():
        if isinstance(sectors, list):
            for s in sectors:
                if s in highlighted:
                    category_colors[category] = highlighted[s]
                    break
        else:
            if sectors in highlighted:
                category_colors[category] = highlighted[sectors]

    # === Title ===
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])
    merge_end_col = min(start_col + title_size, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)

    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = matrix_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # === Write matrix with formatting ===
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            col_name = df.columns[c_idx - 1] if (r_idx > 5 and c_idx > 0 and c_idx - 1 < len(df.columns)) else None

            # Header or index
            if r_idx == 5 or c_idx == 0:
                cell.fill = light_blue
                cell.font = bold_font
            # Column-based background color (below header)
            elif col_name in category_colors:
                fill = PatternFill(start_color=category_colors[col_name], end_color=category_colors[col_name], fill_type="solid")
                cell.fill = fill

            cell.border = black_border
            cell.alignment = center_align

    # === Add separator column ===
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)
    return sep_col + 1


def package_print_embed_plot_option(df_4graph, varname, first_year, last_year, year_range, countries,
                        ICTsectors, ICTcategories, highlighted, cagr_title, stacked_shares_title,
                        end_years_title, xlsx_filename, worksheet_name, start_row, ICT,
                        embed_or_plot):
    
    title_size=6
    # fig 1:  CAGR 
    cagr = clc_cagr(df_4graph, first_year, last_year, varname) 
    if embed_or_plot>0:
        # fig1: plot output CAGR
        plot_cagr(cagr, cagr_title)

    varname_shares, ICT_varname_shares = get_share(df_4graph, varname, first_year, last_year, ICTsectors)
    varname_ICT_share_category = get_share_by_category(ICT_varname_shares, varname, ICTcategories, desired_order)
    if embed_or_plot>0:
        # fig2B: stacked output share
        #this is the average of each category - stacked. 
        #change the following: GDP_ICT_share_category alreadyby category
        _ = plot_stacked_shares(varname_ICT_share_category, varname, ICTcategories, desired_order, stacked_shares_title, highlighted)

    # fig 2C: end years comparison compare first_year with last_year, stacked
    if embed_or_plot>0:
        plot_share_compare_first_last_year(varname_shares, varname, first_year, last_year, end_years_title)
                                            

    # print data to excel (if I take it out of the if statement it will print every time it runs)
    if (embed_or_plot==0) or (embed_or_plot==2):
        # next 10 lines: previousely the function "package print shares to excel"
        start_col = 1
        start_col = create_excel_file_with_title(worksheet_name, xlsx_filename )
        for year in year_range:
            for country in countries: 
                start_col = append_styled_matrix_to_excel(df_4graph[(df_4graph.country==country) & (df_4graph.year==year)], 
                                                          varname, worksheet_name, start_col, xlsx_filename, highlighted, title_size )
                 
        
        start_col = append_styled_matrix_to_excel(varname_shares, varname+'_shares', worksheet_name, start_col, filename=xlsx_filename, highlighted_sectors=highlighted, title_size=title_size)
        start_col = append_styled_matrix_to_excel(ICT_varname_shares, ICT+varname+' shares', worksheet_name, start_col, filename=xlsx_filename, highlighted_sectors=highlighted, title_size=title_size)

        start_col = append_styled_matrix_by_category_to_excel(varname_ICT_share_category, xlsx_filename, worksheet_name, varname+ICT+' by category',start_col, ICTcategories, highlighted, title_size=title_size)
        #                                                     (df,                     filename,      worksheet_name,  matrix_name,              start_col, ICTcategories, highlighted, title_size)

        # embed plots to excel
        col_letter = get_column_letter(start_col)
         
        # === embed1. Create both plots and save to in-memory buffers ===
        # CAGR plot
        fig1 = plot_cagr(cagr, cagr_title)
        buf1 = BytesIO()
        fig1.savefig(buf1, format='png', bbox_inches='tight', dpi=200)
        buf1.seek(0)

        # Stacked shares plot
        fig2 = plot_stacked_shares(varname_ICT_share_category, varname, ICTcategories, desired_order, stacked_shares_title, highlighted)
                                       
                
        
        buf2 = BytesIO()
        fig2.savefig(buf2, format='png', bbox_inches='tight', dpi=200)
        buf2.seek(0)

        # === embed2. Open Excel workbook and worksheet ===
        wb = load_workbook(xlsx_filename)
        ws = wb[worksheet_name]

        # === embed3. Insert both plots ===
        img1 = XLImage(buf1)
        img1.anchor = f"{col_letter}{start_row}"
        ws.add_image(img1)

        img2 = XLImage(buf2)
        img2.anchor = f"{col_letter}{start_row+30}"
        ws.add_image(img2)

        # === embed4. Save the Excel workbook ===
        wb.save(xlsx_filename)
        print(f"✅ Two plots were embedded into '{worksheet_name}' of '{xlsx_filename}'.")


####################################################         version A                 ######################################################
############   plots A version    ###################
# fig 3
def get_one_year_value(df, year, forward_or_backward, sector_list, value_column):
    if forward_or_backward == 'backward':
        col = 'buying sector'
    else:
        col = 'selling sector'
    name = value_column.split()[0]

    #first slicing 
    one_year_impact = df[
        (df['year'] == year) & 
        (df[col].isin(sector_list))       
    ][['country', 'year', 'selling sector', 'buying sector', value_column, f'national {name}']]
    # division
    one_year_impact[value_column] = ( one_year_impact[value_column] / one_year_impact[f'national {name}'] )
    one_year_impact.drop(columns=[f'national {name}'], inplace=True)

    one_year_impact_grouped = one_year_impact.groupby(['country', 'year'], as_index=False)[value_column].sum()
    one_year_impact_grouped = one_year_impact_grouped.sort_values(by=f'{name} impact total', ascending=False)
    return one_year_impact_grouped


def plot_GDPimpact_side_by_side(
    first_year_backwards, last_year_backwards,
    first_year_forwards, last_year_forwards,
    value_column, graph_title):
    width = 0.35
    fig, axes = plt.subplots(1, 2, figsize=(16, 6), sharey=True)

    # --- Panel 1: Backward Linkage ---
    ax = axes[0]
    last_year_backwards = last_year_backwards.sort_values(by=value_column, ascending=False)
    countries = last_year_backwards['country'].values

    first_year_backwards = first_year_backwards.set_index('country').loc[countries].reset_index()
    last_year_backwards = last_year_backwards.reset_index(drop=True)

    x = np.arange(len(countries))

    # Define colors
    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_backwards[value_column], width,
                   label=f"{first_year_backwards['year'].iloc[0]}", color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_backwards[value_column], width,
                   label=f"{last_year_backwards['year'].iloc[0]}", color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Backward Linkage')
    ax.set_xlabel('Country')
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)
    ax.legend()

    # --- Panel 2: Forward Linkage ---
    ax = axes[1]
    last_year_forwards = last_year_forwards.sort_values(by=value_column, ascending=False)
    countries = last_year_forwards['country'].values

    first_year_forwards = first_year_forwards.set_index('country').loc[countries].reset_index()
    last_year_forwards = last_year_forwards.reset_index(drop=True)

    x = np.arange(len(countries))

    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_forwards[value_column], width,
                   label=f"{first_year_forwards['year'].iloc[0]}", color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_forwards[value_column], width,
                   label=f"{last_year_forwards['year'].iloc[0]}", color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Forward Linkage')
    ax.set_xlabel('Country')
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)
    ax.legend()

    fig.supylabel(value_column)
    fig.suptitle(f'{graph_title}, {value_column}: Comparison of {first_year_backwards["year"].iloc[0]} and {last_year_backwards["year"].iloc[0]}')
    plt.tight_layout()
    plt.show()


def plot_GDPimpact_top_bottom(
    first_year_backwards, last_year_backwards,
    first_year_forwards, last_year_forwards,
    value_column, graph_title):

    width = 0.35
    fig, axes = plt.subplots(2, 1, figsize=(12, 10), sharex=False)

    # --- Panel 1: Backward Linkage ---
    ax = axes[0]
    last_year_backwards = last_year_backwards.sort_values(by=value_column, ascending=False)
    countries = last_year_backwards['country'].values

    first_year_backwards = first_year_backwards.set_index('country').loc[countries].reset_index()
    last_year_backwards = last_year_backwards.reset_index(drop=True)

    x = np.arange(len(countries))
    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_backwards[value_column], width, color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_backwards[value_column], width, color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Backward Linkage')
    ax.set_ylabel(value_column)
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)

    # Custom legend (light blue = first year, dark blue = last year)
    custom_legend = [
        Patch(facecolor='skyblue', label=str(first_year_backwards['year'].iloc[0])),
        Patch(facecolor='navy', label=str(last_year_backwards['year'].iloc[0]))
    ]
    ax.legend(handles=custom_legend)

    # --- Panel 2: Forward Linkage ---
    ax = axes[1]
    last_year_forwards = last_year_forwards.sort_values(by=value_column, ascending=False)
    countries = last_year_forwards['country'].values

    first_year_forwards = first_year_forwards.set_index('country').loc[countries].reset_index()
    last_year_forwards = last_year_forwards.reset_index(drop=True)

    x = np.arange(len(countries))
    colors_first = ['lightgreen' if c == 'CAN' else 'skyblue' for c in countries]
    colors_last = ['darkgreen' if c == 'CAN' else 'navy' for c in countries]

    bars1 = ax.bar(x - width/2, first_year_forwards[value_column], width, color=colors_first)
    bars2 = ax.bar(x + width/2, last_year_forwards[value_column], width, color=colors_last)

    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)

    ax.set_title('Forward Linkage')
    ax.set_ylabel(value_column)
    ax.set_xticks(x)
    ax.set_xticklabels(countries, rotation=45)
    ax.legend(handles=custom_legend)

    fig.suptitle(f'{graph_title}, {value_column}: Comparison of {first_year_backwards["year"].iloc[0]} and {last_year_backwards["year"].iloc[0]}')
    plt.tight_layout()
    plt.show()



# fig. 3 for ICT categories
def plot_GDPimpact_wrapper(dfGDPimpact, first_year, last_year, sector_list, value_column, sector_label):
    #this function uses the plot function above to plot 4 different categories of the ICT sector
    # Extract data for backward and forward linkages
    first_year_backward = get_one_year_value(dfGDPimpact, first_year, 'backward', sector_list, value_column)
    last_year_backward = get_one_year_value(dfGDPimpact, last_year, 'backward', sector_list, value_column)
    first_year_forward = get_one_year_value(dfGDPimpact, first_year, 'forward', sector_list, value_column)
    last_year_forward = get_one_year_value(dfGDPimpact, last_year, 'forward', sector_list, value_column)

    # Plot using existing plotting function
    plot_GDPimpact_top_bottom(first_year_backward, last_year_backward,
                              first_year_forward, last_year_forward,
                             'GDP impact total', sector_label)
    # Optionally set a custom plot title
    #plt.suptitle(f'{value_column} for {sector_label}', fontsize=14)
    #plt.tight_layout()
    #plt.show()


# fig. 4 GDP impact backward forward stacked
def plot_stacked_ict_impact(backward_df, forward_df, year, value_col, title):
    
    # Filter for the given year and reindex
    back = backward_df[backward_df['year'] == year][['country', value_col]].set_index('country')
    fwd = forward_df[forward_df['year'] == year][['country', value_col]].set_index('country')

    # Align and fill missing values
    back, fwd = back.align(fwd, join='outer', fill_value=0)

    # Sort by total impact
    total = back[value_col] + fwd[value_col]
    sorted_countries = total.sort_values(ascending=False).index

    # Sort the data
    back_sorted = back.loc[sorted_countries]
    fwd_sorted = fwd.loc[sorted_countries]

    # Create the bar plot
    fig, ax = plt.subplots(figsize=(12, 6))
    for i, country in enumerate(sorted_countries):
        back_val = back_sorted.loc[country, value_col]
        fwd_val = fwd_sorted.loc[country, value_col]
        total_val = back_val + fwd_val

        # Choose colors
        if country == 'CAN':
            back_color = 'forestgreen'
            fwd_color = 'lightgreen'
        else:
            back_color = 'indigo'
            fwd_color = 'orchid'

        # Plot bars
        ax.bar(i, back_val, color=back_color)
        ax.bar(i, fwd_val, bottom=back_val, color=fwd_color)

        # Add percentage label
        ax.text(i, total_val + 0.01, f'{total_val * 100:.1f}%', ha='center', va='bottom', fontsize=9)

    # Final plot settings
    ax.set_title(f"{title} in {year}", fontsize=14)
    ax.set_ylabel('GDP Impact')
    ax.set_xlabel('Country')
    ax.set_xticks(range(len(sorted_countries)))
    ax.set_xticklabels(sorted_countries, rotation=45)

    # Legend
    from matplotlib.patches import Patch
    legend_handles = [
        Patch(color='indigo', label='Backward (others)'),
        Patch(color='orchid', label='Forward (others)'),
        Patch(color='forestgreen', label='Backward (CAN)'),
        Patch(color='lightgreen', label='Forward (CAN)')
    ]
    ax.legend(handles=legend_handles)

    ax.grid(axis='y', linestyle='--', alpha=0.5)
    plt.tight_layout()
    plt.show()



# fig. 5 GDP impact on Education and Health
# fig. 5
def get_one_year_imapct_on_sector(df, year, forward_or_backward, impacting_sectors, impacted_sectors, value_column):
    if forward_or_backward == 'backward':
        col = 'buying sector'
    else:
        col = 'selling sector'
    #first slicing 
    one_year_impact = df[
        (df['year'] == year) & 
        (df[col].isin(impacting_sectors))       
    ][['country', 'year', 'selling sector', 'buying sector', value_column, 'national GDP']]
    # division
    one_year_impact[value_column] = ( one_year_impact[value_column] / one_year_impact['national GDP'] )
    one_year_impact.drop(columns=['national GDP'], inplace=True)
    one_year_impact = one_year_impact[one_year_impact['buying sector'].isin(impacted_sectors)]
    # grouping
    one_year_impact_grouped = one_year_impact.groupby(['country', 'year'], as_index=False)[value_column].sum()
    one_year_impact_grouped = one_year_impact_grouped.sort_values(by='GDP impact total', ascending=False).reset_index(drop=True)
    
    return one_year_impact_grouped



def plot_impact_with_table(df_first, df_last, value_column, graph_title):
    # Extract and validate years
    unique_years_first = df_first['year'].unique()
    unique_years_last = df_last['year'].unique()

    if len(unique_years_first) != 1:
        raise ValueError(f"df_first contains multiple years: {unique_years_first}")
    if len(unique_years_last) != 1:
        raise ValueError(f"df_last contains multiple years: {unique_years_last}")

    first_year = unique_years_first[0]
    last_year = unique_years_last[0]

    # Merge and sort
    merged = df_last[['country', value_column]].merge(
        df_first[['country', value_column]],
        on='country',
        suffixes=('_last', '_first')
    )
    merged = merged.sort_values(by=f'{value_column}_last', ascending=False).reset_index(drop=True)
    merged['Rank'] = merged.index + 1

    # Convert to percent
    merged[f'{value_column}_first'] *= 100
    merged[f'{value_column}_last'] *= 100

    # Prepare table data with % symbols
    table_data = merged[['Rank', 'country', f'{value_column}_first', f'{value_column}_last']].copy()
    table_data[f'{value_column}_first'] = table_data[f'{value_column}_first'].map(lambda x: f'{x:.2f}%')
    table_data[f'{value_column}_last'] = table_data[f'{value_column}_last'].map(lambda x: f'{x:.2f}%')
    table_vals = table_data.values.tolist()

    column_labels = ['Rank', 'Country', str(first_year), str(last_year)]

    countries = merged['country']
    values_first = merged[f'{value_column}_first_raw'] = merged[f'{value_column}_first']
    values_last = merged[f'{value_column}_last_raw'] = merged[f'{value_column}_last']
    x = np.arange(len(countries))
    bar_width = 0.4

    fig = plt.figure(figsize=(16, 6))
    spec = gridspec.GridSpec(ncols=2, nrows=1, width_ratios=[1, 3], wspace=0.3)

    # Table
    ax_table = fig.add_subplot(spec[0])
    table = ax_table.table(
        cellText=table_vals,
        colLabels=column_labels,
        cellLoc='center',
        loc='center'
    )
    ax_table.set_title('Education\nICT Forward GDP Impact Ranking', pad=30, fontsize=12)
    ax_table.axis('off')

    # Bar chart
    ax_bar = fig.add_subplot(spec[1])
    color_first = ['lightgreen' if c == 'CAN' else 'lightblue' for c in countries]
    color_last = ['darkgreen' if c == 'CAN' else 'blue' for c in countries]

    bars1 = ax_bar.bar(x - bar_width/2, values_first, width=bar_width, color=color_first, label=str(first_year))
    bars2 = ax_bar.bar(x + bar_width/2, values_last, width=bar_width, color=color_last, label=str(last_year))

    # Annotations
    for i in range(len(countries)):
        ax_bar.text(x[i] - bar_width/2, values_first[i] + 0.3, f'{values_first[i]:.2f}%', ha='center', va='bottom', fontsize=8)
        ax_bar.text(x[i] + bar_width/2, values_last[i] + 0.3, f'{values_last[i]:.2f}%', ha='center', va='bottom', fontsize=8)

    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(countries, rotation=45, ha='right')
    ax_bar.set_ylabel('GDP Impact (%)')
    ax_bar.set_title(graph_title)

    max_height = max(max(values_first), max(values_last))
    ax_bar.set_ylim(0, max_height * 1.15)

    dummy1 = plt.Rectangle((0,0),1,1,color='lightblue')
    dummy2 = plt.Rectangle((0,0),1,1,color='blue')
    ax_bar.legend([dummy1, dummy2], [str(first_year), str(last_year)], loc='upper right')

    plt.tight_layout()
    plt.show()





##################################################        functions that calculate        ######################################################
# B version
def pivot_matrix_to_3_columns(m: pd.DataFrame, value: str) -> pd.DataFrame:
    # reset_index() selling sector into a column
    return m.reset_index().melt(id_vars=m.index.name or 'index',
                                var_name='buying sector',
                                value_name=value).rename(columns={m.index.name or 'index': 'selling sector'})
# reset_index() turns the row index (selling sector) into a column
# melt(...) collapses all columns (buying sectors) into one column
# id_vars keeps the selling sector fixed
# var_name='buying sector' names the column holding former column labels
# value_name=value names the numbers column
# rename(...) renames the index column to selling sector

def multipliers_direct_indirect_induced(country, year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels):
    n = len(simple_II_labels)
    #notice that dfEj_by_xj and dfGDPj_by_xj were uploaded from B06 and therefore contain only data years
    Ej_by_xj   = dfEj_by_xj[(dfEj_by_xj['country'] == country) & (dfEj_by_xj['year'] == year)].set_index('sector')['Ej_by_xj']
    GDPj_by_xj = dfGDPj_by_xj[(dfGDPj_by_xj['country'] == country) & (dfGDPj_by_xj['year'] == year)].set_index('sector')['GDPj_by_xj']    
    s2s_mo = dfmo[(dfmo['country'] == country) & (dfmo['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="mo").copy()
    s2s_mh = dfmh[(dfmh['country'] == country) & (dfmh['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="mh").copy()
    s2s_mg = dfmg[(dfmg['country'] == country) & (dfmg['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="mg").copy()
    s2s_moc = dfmoc[(dfmoc['country'] == country) & (dfmoc['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="moc").copy()
    s2s_mhc = dfmhc[(dfmhc['country'] == country) & (dfmhc['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="mhc").copy()
    s2s_mgc = dfmgc[(dfmgc['country'] == country) & (dfmgc['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="mgc").copy()
    # correct pivot: HFCE appears in the middle of them matrix (alphabetically) HFCE in rows and employees_compensation in columns
    s2s_moc = s2s_moc[[c for c in s2s_moc.columns if c != 'HFCE'] + ['HFCE']]
    s2s_mhc = s2s_mhc[[c for c in s2s_mhc.columns if c != 'HFCE'] + ['HFCE']]
    s2s_mgc = s2s_mgc[[c for c in s2s_mgc.columns if c != 'HFCE'] + ['HFCE']]
    # direct  
    direct_o = pd.DataFrame(np.eye(n), index=s2s_mo.index, columns=s2s_mo.columns)
    direct_h = pd.DataFrame(np.zeros((n, n)), index=Ej_by_xj.drop("HFCE").index, columns=Ej_by_xj.drop("HFCE").index)
    np.fill_diagonal(direct_h.values, Ej_by_xj.values)
    direct_g = pd.DataFrame(np.zeros((n, n)), index=GDPj_by_xj.drop("HFCE").index, columns=GDPj_by_xj.drop("HFCE").index)
    np.fill_diagonal(direct_g.values, GDPj_by_xj.values)
    #indirect
    indirect_o = (s2s_mo - direct_o).copy()
    #Ej_by_xj*L_minus_I = s2s_mh-Ej_by_xj
    indirect_h  = (s2s_mh - direct_h).copy()
    #GDPj_by_xj*L_minus_I = s2s_mg-GDPj_by_xj
    indirect_g  = (s2s_mg - direct_g).copy()
    #induced
    #first remove HFCE row from s2s_moc, s2s_mhc, s2s_mgc
    induced_o = (s2s_moc.drop('employees_compensation', axis=0).drop('HFCE', axis=1) - s2s_mo).copy()
    induced_h = (s2s_mhc.drop('employees_compensation', axis=0).drop('HFCE', axis=1) - s2s_mh).copy()
    induced_g = (s2s_mgc.drop('employees_compensation', axis=0).drop('HFCE', axis=1) - s2s_mg).copy()
   

    dfdirect_o = pivot_matrix_to_3_columns(direct_o, 'direct_o')
    dfindirect_o = pivot_matrix_to_3_columns(indirect_o, 'indirect_o')
    dfinduced_o = pivot_matrix_to_3_columns(induced_o, 'induced_o')
    dfdirect_h = pivot_matrix_to_3_columns(direct_h, 'direct_h')
    dfindirect_h = pivot_matrix_to_3_columns(indirect_h, 'indirect_h')
    dfinduced_h = pivot_matrix_to_3_columns(induced_h, 'induced_h')
    dfdirect_g = pivot_matrix_to_3_columns(direct_g, 'direct_g')
    dfindirect_g = pivot_matrix_to_3_columns(indirect_g, 'indirect_g')
    dfinduced_g = pivot_matrix_to_3_columns(induced_g, 'induced_g')
    
    multipliers1country1year = dfdirect_o.copy()
    multipliers1country1year = multipliers1country1year.merge(dfindirect_o, on=['selling sector', 'buying sector'], how='left')
    multipliers1country1year = multipliers1country1year.merge(dfinduced_o, on=['selling sector', 'buying sector'], how='left')          
    multipliers1country1year = multipliers1country1year.merge(dfdirect_h, on=['selling sector', 'buying sector'], how='left')
    multipliers1country1year = multipliers1country1year.merge(dfindirect_h, on=['selling sector', 'buying sector'], how='left')
    multipliers1country1year = multipliers1country1year.merge(dfinduced_h, on=['selling sector', 'buying sector'], how='left')  
    multipliers1country1year = multipliers1country1year.merge(dfdirect_g, on=['selling sector', 'buying sector'], how='left')
    multipliers1country1year = multipliers1country1year.merge(dfindirect_g, on=['selling sector', 'buying sector'], how='left')          
    multipliers1country1year = multipliers1country1year.merge(dfinduced_g, on=['selling sector', 'buying sector'], how='left')
    multipliers1country1year["country"] = country
    multipliers1country1year["year"] = year     

  
    return multipliers1country1year   


def multipliers2prediction(s2s_m, fdf_year2, column_name):
    predicted_output_year2_np  = np.round(s2s_m.to_numpy() @ fdf_year2.values.reshape(-1, 1), 1)
    
    predicted_output_year2 = pd.DataFrame(predicted_output_year2_np, index=s2s_m.index, columns=[column_name])
    
    return predicted_output_year2
# to delete
def scale_df_by_series(direct_o: pd.DataFrame, fcdf: pd.Series) -> pd.DataFrame:
    
    return direct_o[fcdf.index].mul(fcdf, axis=1)


#################################################                 check               #############################################
def check_L_Lc(country, year, dfTc, dfoutput, df8, df9):
        HFCE_col_name = 'HFCE'
        f8_col_name = "8 final demand"
        f9_col_name = "9 final demand"
        GDP_col_name = 'GDP'
        Tctemp = (
            dfTc[(dfTc.country == country) & (dfTc.year == year)]
            .pivot(index="selling_sector", columns="buying_sector", values="Tc")
        )
        x = dfoutput[(dfoutput.country == country) & (dfoutput.year == year)].set_index('sector')['output']
        f8temp = df8[(df8.country == country) & (df8.year == year)].set_index('sector')[f8_col_name]
        f8temp.rename(index={"employees_compensation":"HFCE"},inplace=True) 
        f9temp = df9[(df9.country == country) & (df9.year == year)].set_index('sector')[f9_col_name]
        f9temp.rename(index={"employees_compensation":"HFCE"},inplace=True) 
        Tctemp = Tctemp[[c for c in Tctemp.columns if c != 'HFCE'] + ['HFCE']] #put 'HFCE' at the end
        Lctemp,_ = clc_L(Tctemp)
        xcheck = multipliers2prediction(Lctemp, f8temp,'output') #it is important to use this instead of @ because when the row names don't agree the multiplication is inaccurate  
        #Lc shoudl be multilied by f8
        Ltemp,_ = clc_L(Tctemp.iloc[:-1,:-1])
        xcheck2 = Ltemp@f9temp.iloc[:-1]
        print('        xcheck   ',' xcheck2', '   x dfoutput')
        print(pd.concat([xcheck.iloc[:-1].round(), xcheck2.round(), x.iloc[:-1].round() ], axis=1))


def check_impacts():
    # the following comfirms that the impacts are correct
    #currently I calculate impacts by multiplying by f8 and not f9
    # should ask consultant if this is common practice

    #continue from here to make impact graphs

    country = "CAN"
    year = 2012
    output = dfoutput[(dfoutput['country'] == country) & (dfoutput['year'] == year)].set_index("sector")['output'].copy()

    multipliers_CAN2012 = multipliers_direct_indirect_induced(country, year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
    mul_direct = multipliers_CAN2012[["selling sector", "buying sector", 'direct_o']].copy().pivot_table(index=['selling sector'], columns='buying sector',
        values='direct_o')
    mul_indirect = multipliers_CAN2012[["selling sector", "buying sector", 'indirect_o']].copy().pivot_table(index=["selling sector"], columns="buying sector", 
                                                                                                        values='indirect_o') 
    mul_induced = multipliers_CAN2012[["selling sector", "buying sector", 'induced_o']].copy().pivot_table(index=["selling sector"], columns="buying sector", 
                                                                                                        values='induced_o')

    f8_CAN2012 = df8[(df8['country'] == country) & (df8['year'] == year)].set_index("sector")['8 final demand'].copy()
    f9_CAN2012 = df9[(df9['country'] == country) & (df9['year'] == year)].set_index("sector")['9 final demand'].copy()
    #f8_CAN2012.rename(index={"employees_compensation":"HFCE"},inplace=True) to delete
    #f9_CAN2012.rename(index={"employees_compensation":"HFCE"},inplace=True) te delete
    _, _, im_direct_o, im_indirect_o, im_induced_o = \
        get_vector_impacts(country, year, multipliers_CAN2012, f8_CAN2012, f9_CAN2012, "output",'f8')
    #country_temp, year_temp,im_direct_h, im_indirect_h, im_induced_h = get_impacts(country, year, multipliers_CAN2012, f8_CAN2012, f9_CAN2012, "Employment")
    #country_temp, year_temp,im_direct_g, im_indirect_g, im_induced_g = get_impacts(country, year, multipliers_CAN2012, f8_CAN2012, f9_CAN2012, "GDP")
    x_tot_CAN2012 = im_direct_o + im_indirect_o + im_induced_o
    #x_tot_CAN2012_original = dfoutput[(dfoutput['country']==country) & (dfoutput['year']==year)].set_index('sector')['output']
    m_tot = mul_direct+mul_indirect    
    output_mul = multipliers2prediction(mul_direct + mul_indirect + mul_induced, f8_CAN2012.drop("HFCE"),'output')
    s2s_moc = dfmoc[(dfmoc['country'] == country) & (dfmoc['year'] == year)].pivot(index="selling_sector",columns="buying_sector",values="moc").copy()
    s2s_moc = s2s_moc[[c for c in s2s_moc.columns if c != 'HFCE'] + ['HFCE']]
    output_moc = multipliers2prediction(s2s_moc, f8_CAN2012, "output")
    print('                 output   ',' s2s_moc@f8','      output_mul',' impacts direct+indirect+induced')
    print(pd.concat([output.iloc[:-1].round(), output_moc.round() , output_mul, x_tot_CAN2012], axis=1))


#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@                    main                  @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# upload gdp
dfgdp_worldbank = pd.read_csv("Bench_predictions_B/A04_gdp_ARIMAgdp_currentUSD04.csv")
worldbank_gdp_col_name = "gdp total world bank"
dfgdp_worldbank.rename(columns={"Unnamed: 0": "year","gdp total": worldbank_gdp_col_name}, inplace=True)
dfgdp_worldbank.iloc[:, 1:] = dfgdp_worldbank.iloc[:, 1:] * 10**(-6)
dfgdp_worldbank['year'] = dfgdp_worldbank['year'].astype(int)
dfgdp_worldbank = dfgdp_worldbank.set_index('year')


# upload E
dfE = pd.read_csv("Bench_predictions_B/A05_Esectors_from_Etot05.csv")
dfE.rename(columns={"E": "Employment"}, inplace=True)
# dfE already has data until 2040 - base extrapolation already done.

# for base extrapolation
dfTcbase = pd.read_csv("Bench_predictions_B/B071_Tc_base_1years.csv")
dfHFCE = pd.read_csv("Bench_predictions_B/B072_dfHFCE_data_and_extrap.csv")
df8    = pd.read_csv("Bench_predictions_B/B072_df8_data_and_extrap.csv")
df9    = pd.read_csv("Bench_predictions_B/B072_df9_data_and_extrap.csv")
dfGDP  = pd.read_csv("Bench_predictions_B/B072_dfGDP_data_and_extrap.csv")
dfoutput = pd.read_csv("Bench_predictions_B/B072_dfoutput_data_and_extrap.csv")
#dfGDPj_by_xj = pd.read_csv("Bench_predictions_B/B072_dfGDPj_by_xj_data_and_extrap.csv")
#dfEj_by_xj = pd.read_csv("Bench_predictions_B/B072_dfEj_by_xj_data_and_extrap.csv") - this file doesn't exit. If I want it I should make it copy from dfGDPj_by_xj in B07
#upload from B06 and not B072 because I want only data years (for impact calculation)
dfGDPj_by_xj = pd.read_csv("Bench_predictions_B/B06_dfGDPj_by_xj.csv") 
dfEj_by_xj = pd.read_csv("Bench_predictions_B/B06_dfEj_by_xj.csv")

#multipliers, pivoted to long form 
dfmo = pd.read_csv("Bench_predictions_B/B06_dfmo.csv")
dfmoc = pd.read_csv("Bench_predictions_B/B06_dfmoc.csv")
dfmh = pd.read_csv("Bench_predictions_B/B06_dfmh.csv")
dfmhc = pd.read_csv("Bench_predictions_B/B06_dfmhc.csv")
dfmg = pd.read_csv("Bench_predictions_B/B06_dfmg.csv")
dfmgc = pd.read_csv("Bench_predictions_B/B06_dfmgc.csv")        
dfTc = pd.read_csv("Bench_predictions_B/B06_dfTc.csv") # different than Tbase


########################################                           parameters                       ##################################################
start_time = time.time()
print("working directory of B12_graphs4-6.py is: ",os.getcwd())  # Print the current working directory

table_type = 'TTL' #or'DOM'   
OECD_path = "../Data/" # windows style: r".\\"
if table_type == 'DOM':
    output_filename = '/mnt/c/NavitComputer24/2024_NES/Economics/Textbook_EIA/OECD_salaries/EIA_matrices.xlsx'
    final_demand_columns = ['HFCE',	'NPISH', 'GGFC',	'GFCF',	'INVNT', 'CONS_NONRES', 'EXPO']
elif table_type == 'TTL':
    output_filename = '/mnt/c/NavitComputer24/2024_NES/Economics/Textbook_EIA/OECD_salaries/EIA_TTL_matrices.xlsx'
    final_demand_columns = ['HFCE', 'NPISH', 'GGFC', 'GFCF', 'INVNT', 'DPABR', 'CONS_NONRES', 'EXPO', 'IMPO']

# possible
# make a function out of B07_base_sectros so that I can run it with parameters (n_years_for_base for example)

# future years
#max_future_year = dfoutput_tot.year.unique().max()
#year_range_future = [int(year) for year in range(int(last_year+1), max_future_year+1)]

#report_title = f'ICT sectors, {last_year}'
ICTcategories = {'ICT - Manufacturing': 'C26',
                'ICT - Wholesaling': 'G',
                'ICT - Software and computer services': ['J58T60', 'J62_63', 'M'],  
                'ICT - Communications services': 'J61'}
ICTsectors = ['C26', 'G', 'J58T60', 'J62_63', 'M', 'J61']

country_names = ['Canada', 'The United States', 'Great Britain', 'France', 'Germany', 'Italiy', 'Japan']
countries = ['CAN', 'USA', 'GBR', 'FRA', 'DEU', 'ITA', 'JPN'] # 'CHN' is not available in OECD, but it is in OECDadditional
country_map = dict(zip(countries, country_names))

currency_exchange_type = 'EXCH' #'EXCH' or 'PPP'

simple_II_labels = ['A01_02', 'A03', 'B05_06', 'B07_08', 'B09', 'C10T12', 'C13T15', 'C16', 'C17_18', 'C19', 'C20', 'C21', 'C22', 'C23', 'C24', 
                 'C25', 'C26', 'C27', 'C28', 'C29', 'C30', 'C31T33', 'D', 'E', 'F', 'G', 'H49', 'H50', 'H51', 'H52', 'H53', 'I', 'J58T60', 'J61',
                  'J62_63', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']



first_year = 2015 
last_year = 2024  
year_range = [int(year) for year in range(int(first_year), int(last_year) + 1)]
data_years = [int(year) for year in range(int(1995), int(2020) + 1)]
base_year = 2020
dfoutput_4graph = dfoutput[(dfoutput.year >= first_year) & (dfoutput.year <= last_year) ].copy()
dfGDP_4graph = dfGDP[(dfGDP.year >= first_year) & (dfGDP.year <= last_year) ].copy()

highlighted = {
    "C26": "90EE90",   # light green
    "G": "CD7F32",     # bronze
    "J61": "6699CC",   # blue-gray
    "J58T60": "FFFACD",
    "J62_63": "FFFACD",
    "M": "FFFACD"
}

##########################################             Benchmark  plots  1-2          ######################################################

# graphs 1 and 2 for output
if 0:
    graphnumber=1
    varname = 'output'
    ICT = 'ICT'
    cagr_title = f'Average {varname} CAGR for {ICT} sectors ({first_year}–{last_year})'
    end_years_title = f'{ICT} {varname} {first_year} and {last_year} Share by Country'
    stacked_shares_title = f'Stacked Average {ICT} {varname} Share by Country, {first_year}-{last_year}'
    xlsx_filename = f"Bench_predictions_B/B12_graph{graphnumber}_data {first_year}-{last_year}.xlsx"
    worksheet_name = f"{varname} shares {first_year}-{last_year}"
    embed_or_plot = 2 #0: embed, 1: plot, 2: both
    start_row=5
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling',
                        'ICT - Software and computer services', 'ICT - Communications services']


    package_print_embed_plot_option(dfoutput_4graph, varname, first_year, last_year, year_range, countries,
                            ICTsectors,
                            ICTcategories,
                            highlighted,
                            cagr_title,
                            stacked_shares_title,
                            end_years_title,
                            xlsx_filename,
                            worksheet_name,
                            start_row,
                            ICT,
                            embed_or_plot
    )
 # graphs 1 and 2 for GDP

if 0:
   
    graphnumber=1
    varname = 'GDP'
    ICT = 'ICT'
    cagr_title = f'Average {varname} CAGR for ICT sectors ({first_year}–{last_year})'
    end_years_title = f'{ICT} {varname} {first_year} and {last_year} Share by Country'
    stacked_shares_title = f'Stacked Average ICT {varname} Share by Country, {first_year}-{last_year}'
    xlsx_filename = f"Bench_predictions_B/B12_graph{graphnumber}_data {first_year}-{last_year}.xlsx"
    worksheet_name = f"{varname} shares {first_year}-{last_year}"
    embed_or_plot = 0 #0: embed, 1: plot, 2: both
    start_row=5
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling',
                        'ICT - Software and computer services', 'ICT - Communications services']


    package_print_embed_plot_option(dfGDP_4graph, varname, first_year, last_year, year_range, countries,
                            ICTsectors,
                            ICTcategories,
                            highlighted,
                            cagr_title,
                            stacked_shares_title,
                            end_years_title,
                            xlsx_filename,
                            worksheet_name,
                            start_row,
                            ICT,
                            embed_or_plot
    )

#one file for both GDP and output
if 0:
    graphnumber=1
    xlsx_filename = f"Bench_predictions_B/B12_graph{graphnumber}_data_{first_year}-{last_year}.xlsx"
    ICT = 'ICT'
    embed_or_plot = 0 #0: embed, 1: plot, 2: both
    start_row=5
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling',
                        'ICT - Software and computer services', 'ICT - Communications services']
    
    varname = 'output'
    cagr_title = f'Average {varname} CAGR for {ICT} sectors ({first_year}–{last_year})'
    end_years_title = f'{ICT} {varname} {first_year} and {last_year} Share by Country'
    stacked_shares_title = f'Stacked Average {ICT} {varname} Share by Country, {first_year}-{last_year}'
    worksheet_name = f"{varname} shares {first_year}-{last_year}"
    
    package_print_embed_plot_option(dfoutput_4graph, varname, first_year, last_year, year_range, countries,
                            ICTsectors, ICTcategories, highlighted, cagr_title, stacked_shares_title,
                            end_years_title, xlsx_filename, worksheet_name, start_row, ICT,
                            embed_or_plot
    )


    varname = 'GDP'
    cagr_title = f'Average {varname} CAGR for ICT sectors ({first_year}–{last_year})'
    end_years_title = f'{ICT} {varname} {first_year} and {last_year} Share by Country'
    stacked_shares_title = f'Stacked Average ICT {varname} Share by Country, {first_year}-{last_year}'
    worksheet_name = f"{varname} shares {first_year}-{last_year}"
    
    package_print_embed_plot_option(dfGDP_4graph, varname, first_year, last_year, year_range, countries,
                            ICTsectors, ICTcategories, highlighted, cagr_title, stacked_shares_title,
                            end_years_title, xlsx_filename, worksheet_name, start_row, ICT,
                            embed_or_plot
    )

SHRED=0
if SHRED:
    df_can_2030 = dfoutput[(dfoutput['country']=="CAN") &
                           (dfoutput['year']==2030)].iloc[:-1,:]

    plt.figure(figsize=(18,4))   # wider
    plt.plot(df_can_2030['sector'], df_can_2030['output'], marker='o')

    plt.xticks(rotation=75, ha='right')  # strong rotation + right align

    plt.xlabel("Sector")
    plt.ylabel("output, Milions, current USD")
    plt.title("Output by sector for Canada in 2030")
    plt.tight_layout()
    plt.show()


    df_can_2030_GDP = dfGDP[(dfoutput['country']=="CAN") &
                           (dfoutput['year']==2030)].iloc[:-1,:]

    plt.figure(figsize=(18,4))   # wider
    plt.plot(df_can_2030_GDP['sector'], df_can_2030_GDP['GDP'], marker='o')

    plt.xticks(rotation=75, ha='right')  # strong rotation + right align

    plt.xlabel("Sector")
    plt.ylabel("GDP, Milions, current USD")
    plt.title("GDP by sector for Canada in 2030")
    plt.tight_layout()
    plt.show()

print('B12 graphs 1 and 2 are checked')




    
if 1:
    if 0:
        check_impacts()
    #plot backward GDP impact % share + excel
    
    dftemp,  dfGDPimpact = pd.DataFrame(), pd.DataFrame()
    #multipliers can be calculated only for data years
    for country in countries:
        for year in year_range:
            #when making f8 and f9: sector should become index
            f8 = df8[(df8['country'] == country) & (df8['year'] == year)].set_index("sector")['8 final demand'].copy()
            f9 = df9[(df9['country'] == country) & (df9['year'] == year)].set_index("sector")['9 final demand'].copy()
            if year in data_years:
                # multipliers_direc_indirect_induced is only for 1 country and 1 year!
                multipliers = multipliers_direct_indirect_induced(country, year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
            else:
                multipliers = multipliers_direct_indirect_induced(country, base_year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
            _, _, im_direct_g, im_indirect_g, im_induced_g = \
                get_vector_impacts(country, year, multipliers, f8, f9, "GDP",'f9')
            # for calculation use f9
            dftemp = im_direct_g.rename(columns={"GDP": "direct GDP"}).copy()
            dftemp = dftemp.merge(im_indirect_g.rename(columns={"GDP": "indirect GDP"}), left_index=True, right_index=True, how="left")
            dftemp = dftemp.merge(im_induced_g.rename(columns={"GDP": "induced GDP"}), left_index=True, right_index=True, how="left")
            dftemp["sector"] = dftemp.index
            dftemp["country"] = country
            dftemp["year"] = year
            dftemp["backward_forward"] = "backward"
            dftemp = dftemp[["country", "year", "sector", "backward_forward", "direct GDP", "indirect GDP", "induced GDP"]]
            dftemp = dftemp.reset_index(drop=True)
            dfGDPimpact = pd.concat([dfGDPimpact, dftemp], ignore_index=True)

    dfGDPimpact["sum GDP impact"] = dfGDPimpact.loc[:,"direct GDP"]+ dfGDPimpact.loc[:, "indirect GDP"] + dfGDPimpact.loc[:, "induced GDP"]
    dfGDPimpact["direct+indirect GDP impact"] = dfGDPimpact.loc[:,"direct GDP"]+ dfGDPimpact.loc[:, "indirect GDP"]
#GDP impact instead of GDP
#dfGDPimpact instead of dfGDP_4graph

    graphnumber=3
    varname =  "sum GDP impact" #"direct+indirect GDP impact"
    ICT = 'ICT'
    cagr_title = f'Average {varname} CAGR for ICT sectors ({first_year}–{last_year})'
    end_years_title = f'{ICT} {varname} {first_year} and {last_year} Share by Country'
    stacked_shares_title = f'Stacked Average ICT {varname} Share by Country, {first_year}-{last_year}'
    xlsx_filename = f"Bench_predictions_B/B12_graph{graphnumber}_data {first_year}-{last_year}.xlsx"
    worksheet_name = f"{varname} shares {first_year}-{last_year}"
    embed_or_plot = 0 #0: embed, 1: plot, 2: both
    start_row=5
    desired_order = ['ICT - Manufacturing', 'ICT - Wholesaling',
                        'ICT - Software and computer services', 'ICT - Communications services']


    package_print_embed_plot_option(dfGDPimpact, varname, first_year, last_year, year_range, countries,
                            ICTsectors,
                            ICTcategories,
                            highlighted,
                            cagr_title,
                            stacked_shares_title,
                            end_years_title,
                            xlsx_filename,
                            worksheet_name,
                            start_row,
                            ICT,
                            embed_or_plot
    )


##########################################             Benchmark  plots version A         ######################################################
if 0:
    # fig 5. Education graphs
    ICT_on_Education_first_year_forward_impact = get_one_year_imapct_on_sector(dfGDPimpact, first_year, 'forward', ICTsectors, ['P'], 'GDP impact total')
    ICT_on_Education_last_year_forward_impact = get_one_year_imapct_on_sector(dfGDPimpact, last_year, 'forward', ICTsectors, ['P'], 'GDP impact total')

    plot_impact_with_table(ICT_on_Education_first_year_forward_impact, ICT_on_Education_last_year_forward_impact, 'GDP impact total', 'GDP Impact Total ICT Sectors Forward Impact on Education')

    # fig 6. Health graphs
    ICT_on_Education_first_year_forward_impact = get_one_year_imapct_on_sector(dfGDPimpact, first_year, 'forward', ICTsectors, ['Q'], 'GDP impact total')
    ICT_on_Education_last_year_forward_impact = get_one_year_imapct_on_sector(dfGDPimpact, last_year, 'forward', ICTsectors, ['Q'], 'GDP impact total')
    plot_impact_with_table(ICT_on_Education_first_year_forward_impact, ICT_on_Education_last_year_forward_impact, 'GDP impact total', 'GDP Impact Total ICT Sectors Forward Impact on Health')


# fig 7: Employment stacked backward forward
# graph 4 GDP stacked backward forward
if 0:
    ICT_last_year_backward_impact = get_one_year_value(dfEimpact, last_year,'backward', ICTsectors, 'Employment impact total')
    ICT_last_year_forward_impact = get_one_year_value(dfEimpact, last_year,'forward', ICTsectors, 'Employment impact total')
    plot_stacked_ict_impact(ICT_last_year_backward_impact, ICT_last_year_forward_impact, year, 'Employment impact total', 'ICT Employment Impact')


# graph 4 GDP stacked backward forward
if 0:
    ICT_first_year_backward_impact= get_one_year_value(dfGDPimpact, first_year,'backward', ICTsectors, 'GDP impact total')
    ICT_last_year_backward_impact = get_one_year_value(dfGDPimpact, last_year,'backward', ICTsectors, 'GDP impact total')
    ICT_first_year_forward_impact= get_one_year_value(dfGDPimpact, first_year,'forward', ICTsectors, 'GDP impact total')
    ICT_last_year_forward_impact = get_one_year_value(dfGDPimpact, last_year,'forward', ICTsectors, 'GDP impact total')
    plot_stacked_ict_impact(ICT_last_year_backward_impact, ICT_last_year_forward_impact, year, 'GDP impact total', 'ICT GDP Impact')




#####################  backward-forward GDP plot    ######################

'''
if 1:
#    graph 3?
    first_year = 2015
    last_year = 2024
    base_year = 2020
    #multipliers can be calculated only for data years
    if first_year in data_years:
        # multipliers_direc_indirect_induced is only for 1 country and 1 year!
        multipliers_first_year = multipliers_direct_indirect_induced(country, first_year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
    else:
        multipliers_first_year = multipliers_direct_indirect_induced(country, base_year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
    if last_year in data_years:
        multipliers_last_year = multipliers_direct_indirect_induced(country, last_year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
    else:
        multipliers_last_year = multipliers_direct_indirect_induced(country, base_year, dfmo, dfmoc, dfmh, dfmhc, dfmg, dfmgc, dfEj_by_xj, dfGDPj_by_xj, simple_II_labels)
    # when making f8 and f9: sector should become index
     

    f8_first_year = df8[(df8['country'] == country) & (df8['year'] == first_year)].set_index("sector")['8 final demand'].copy()
    f9_first_year = df9[(df9['country'] == country) & (df9['year'] == first_year)].set_index("sector")['9 final demand'].copy()
    f8_last_year   = df8[(df8['country'] == country) & (df8['year'] == last_year)].set_index("sector")['8 final demand'].copy()
    f9_last_year   = df9[(df9['country'] == country) & (df9['year'] == last_year)].set_index("sector")['9 final demand'].copy()
    
    #################################
    # impacts instead of multipliers
    #################################
    #comment abnormality in OECD structure:
    #fdf = OECD.loc[simple_II_labels, final_demand_columns].sum(axis=1)
    #there is what causes closed model to be in accuarete:
    #fcdf_year2 = OECD_year2.loc[simple_II_labels,final_demand_columns[1:]].sum(axis=1)
    #I should take HFCE inside fcdf_year2. 
    #fcdf = OECD.loc[simple_II_labels,final_demand_columns].sum(axis=1)
    #fcdf.loc['employees_compensation'] = 0
    
    # 1 country 1 year
    country_temp, year_temp, im_direct_o, im_indirect_o, im_induced_o = \
    get_vector_impacts(country, start_year, multipliers_start_year, f8_start_year, f9_start_year, "output",'f9')
    country_temp, year_temp,im_direct_h, im_indirect_h, im_induced_h = \
    get_vector_impacts(country, start_year, multipliers_start_year, f8_start_year, f9_start_year, "Employment",'f9')
    country_temp, year_temp,im_direct_g, im_indirect_g, im_induced_g = \
    get_vector_impacts(country, start_year, multipliers_start_year, f8_start_year, f9_start_year, "GDP",'f9')
    
    
    
    
    # back to A version:

    # then go below and do 1 year value differently    
    ICT_start_year_backward_impact= get_one_year_value(dfGDPimpact, start_year,'backward', ICTsectors, 'GDP impact total')
    ICT_end_year_backward_impact = get_one_year_value(dfGDPimpact, end_year,'backward', ICTsectors, 'GDP impact total')
    ICT_start_year_forward_impact= get_one_year_value(dfGDPimpact, start_year,'forward', ICTsectors, 'GDP impact total')
    ICT_end_year_forward_impact = get_one_year_value(dfGDPimpact, end_year,'forward', ICTsectors, 'GDP impact total')
    plot_GDPimpact_top_bottom( ICT_first_year_backward_impact, ICT_last_year_backward_impact,
                            ICT_first_year_forward_impact, ICT_last_year_forward_impact,
                            'GDP impact total','ICT' )
    
    for sector_label, sector_list in ICT_factors.items():
        if isinstance(sector_list, str):
            sector_list = [sector_list]
        plot_GDPimpact_wrapper(dfGDPimpact, first_year, last_year, sector_list, 'GDP impact total', sector_label)
        # plot_GDPimpact_top_bottom is inside it
    '''




end_time = time.time()
print(f"Elapsed time: {(end_time - start_time)/60:.1f} minutes")


