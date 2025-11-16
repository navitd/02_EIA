

import warnings
# Suppress FutureWarnings globally
warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', message=r".*force_all_finite.*")
warnings.filterwarnings("ignore")

import pandas as pd
import matplotlib.pyplot as plt
import sys
from pathlib import Path
import os
# Add the parent directory to sys.path
sys.path.append(str(Path(__file__).resolve().parent.parent / 'EIAfunctions'))
from func_clean_world_bank_data import clean_world_bank_data
from sklearn.metrics import mean_squared_error
import pmdarima as pm

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import openpyxl




def get_worldbank_gdp_data(plot_flag):

    # Define the data directory relative to the script location
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    GDP_filename = os.path.join(SCRIPT_DIR, '..', '..', 'Data', 'World Bank G7 GDP', 
                            'cb6272d5-797d-4d20-970a-1286e5b13605_Data.csv')
    #Data from database: World Development Indicators
    #Last Updated: 07/01/2025

    # Add error handling for file loading

    if not os.path.exists(GDP_filename):
        raise FileNotFoundError(f"Data file not found at: {GDP_filename}")

    rough = pd.read_csv(GDP_filename)
    print(f"Successfully loaded data with shape: {rough.shape}")
        
    worldbank_gdp_data, file_description = clean_world_bank_data(rough)

    print("\n dataframe data:\n")
    print(worldbank_gdp_data.head())

    # Plotting
    if plot_flag:
        countries = data.columns[1:]  # all country columns
        plt.figure(figsize=(10, 6))

        for country in countries:
            plt.plot(data['Time'], data[country], marker = 'o',label=country)
            plt.text(data['Time'].iloc[-1], data[country].iloc[-1], country, 
                    fontsize=9, verticalalignment='center')

        plt.xlabel('Year')
        plt.ylabel(file_description[2])  
        plt.title('file_description[2] of countries')
        plt.grid(True)
        plt.legend()  
        plt.show()

        # Trending
        max_data_years = data.Time.max()
        past_years = 10


    return worldbank_gdp_data


def plot_gdp_forecast(worldbank_gdp_data, forecast_gdp, countries, title):
    
    plt.figure(figsize=(12, 6))

    for country in countries:
        # Plot historical data and get the color assigned
        line_actual, = plt.plot(worldbank_gdp_data.index, worldbank_gdp_data[country], label=f"{country} actual")
        color = line_actual.get_color()
        
        # Plot forecasted data with the same color
        plt.plot(forecast_gdp.index, forecast_gdp[country], linestyle='--', color=color, label=f"{country} forecast")

    plt.xlabel("Year")
    plt.ylabel("GDP (current USD)")
    plt.title(title)
    plt.legend()
    plt.grid(True)
    plt.show()


###############################################  printing to Excel  #####################################################

def create_excel_file_with_title(year: str, filename: str = "output.xlsx") -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = year

    # Styles
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Merge title box over first 4 columns and 3 rows
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=4)
    cell = ws.cell(row=1, column=1)
    cell.value = f"EIA details - {year}"
    cell.fill = green
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = black_border

    wb.save(filename)

    return 1  # Next available column after title box

def append_styled_matrix_to_excel(df, matrix_name, year: str, start_col: int, filename: str = "output.xlsx", title_size=3) -> int:
    # Infer matrix name from variable name if not provided
    if matrix_name is None:
        frame = inspect.currentframe().f_back
        matrix_name = next((name for name, val in frame.f_locals.items() if val is df), "UnnamedMatrix")

    wb = openpyxl.load_workbook(filename)
    if year not in wb.sheetnames:
        raise ValueError(f"Sheet named '{year}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[year]

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Convert DataFrame to rows (including index and header)
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # includes index

    # Green title merged over up to 4 columns
    merge_end_col = min(start_col + title_size, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = matrix_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the matrix below the title
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)

    return sep_col + 1  # Return column to start the next matrix (skip separator too)

def append_styled_series_to_excel(series: pd.Series, series_name, year: str, start_col: int, filename: str = "output.xlsx") -> int:
    # Infer series name from variable name if not provided
    if series_name is None:
        frame = inspect.currentframe().f_back
        series_name = next((name for name, val in frame.f_locals.items() if val is series), "UnnamedSeries")

    wb = openpyxl.load_workbook(filename)
    if year not in wb.sheetnames:
        raise ValueError(f"Sheet named '{year}' does not exist. Create it first using create_excel_file_with_title.")
    ws = wb[year]

    # Convert Series to DataFrame for uniformity
    df = series.to_frame(name=series.name if series.name else "Value")
    rows = list(dataframe_to_rows(df, index=True, header=True))
    n_rows = len(rows)
    n_cols = len(rows[0])  # Should be 2: index and value

    # Styles
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="00C000", end_color="00C000", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Green title merged over up to 2 columns
    merge_end_col = min(start_col + 1, start_col + n_cols - 1)
    if merge_end_col > start_col:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=merge_end_col)
    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = series_name
    title_cell.fill = green
    title_cell.font = bold_font
    title_cell.alignment = center_align

    # Write the series below the title
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row):
            col = start_col + c_idx
            cell = ws.cell(row=r_idx, column=col, value=val)
            if r_idx == 5 or c_idx == 0:  # header or index
                cell.fill = light_blue
                cell.font = bold_font
            cell.border = black_border
            cell.alignment = center_align

    # Add a black separator column
    sep_col = start_col + n_cols
    for r in range(4, 5 + n_rows):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = black_border
        cell.alignment = center_align

    wb.save(filename)
    return sep_col + 1








# 01 upload gdp data
worldbank_gdp_data = get_worldbank_gdp_data(False)
worldbank_gdp_data.rename(columns={'Time': 'year'}, inplace=True   ) #renaming the column
worldbank_gdp_data['year'] = worldbank_gdp_data['year'].astype(int)  #recasting as int
worldbank_gdp_data.set_index('year', inplace=True)                   #setting year as index

# 1. Train–test split (e.g., 80% train, 20% test)
train_test_split = 0.8
train_size = int(len(worldbank_gdp_data) * train_test_split)
train, test = worldbank_gdp_data[:train_size], worldbank_gdp_data[train_size:]

forecast_horizon = 16  # years ahead to predict
forecast_years = list(range(worldbank_gdp_data.index.max() + 1,
                            worldbank_gdp_data.index.max() + forecast_horizon + 1))
countries = ['ITA','JPN','CAN','FRA','DEU','GBR','USA']

forecast_gdp = pd.DataFrame(index=pd.Index(forecast_years, dtype=int), columns=countries)

arima_orders = {}
for country in countries:
    print(f"\nProcessing {country}...")
    
    # Extract series
    train = worldbank_gdp_data[:train_size][country]
    test = worldbank_gdp_data[train_size:][country]
    
    # Systematic grid search for ARIMA(p,d,q)
    # d=1 sets difference y(t)-y(t-1)
    #fixed d:
    if 1:
        d_title = 3
        model = pm.auto_arima(
            train,
            start_p=0, max_p=5,     # search range for AR terms
            start_q=0, max_q=5,     # search range for MA terms
            d=d_title,                 # let model decide via ADF test
            seasonal=False,
            stepwise=False,         # disable heuristic stepwise search
            suppress_warnings=True,
            error_action='ignore',
            information_criterion='aic',  # use AIC for model selection
            n_jobs=-1               # parallel processing for speed
        )

    #automatic d
    if 0:
        d_title = 'automatic'
        model = pm.auto_arima(
            train,
            start_p=0, max_p=5,
            start_q=0, max_q=5,
            d=None,               # <-- let auto_arima choose differencing
            test='adf',           # unit-root test to determine d
            seasonal=False,
            stepwise=False,
            suppress_warnings=True,
            error_action='ignore',
            information_criterion='aic',
            n_jobs=-1
        )
    
    print(f"\n\nSelected ARIMA order for {country}: {model.order}\n")
    
    # Forecast on test to evaluate
    prediction_on_test = model.predict(n_periods=len(test))
    mse = mean_squared_error(test, prediction_on_test)
    print(f"Test MSE for {country}: {mse:.2e}")
    
    # Refit on full series
    model.fit(worldbank_gdp_data[country])
    arima_orders[country] = model.order

    # Forecast 10 years ahead
    forecast_values = model.predict(n_periods=forecast_horizon)
    forecast_values.index = forecast_years
    
    # Assign to column
    forecast_gdp[country] = forecast_values


plot_gdp_forecast(worldbank_gdp_data, forecast_gdp, countries, title=f"ARIMA gdp d={d_title}")

print("\nSelected ARIMA orders for all countries:\n")
for country, order in arima_orders.items():
    print(f"{country}: ARIMA{order}")

# print to excel
if 0:
    gdp = pd.concat([worldbank_gdp_data, forecast_gdp])
    gdp.to_csv("Bench_predictions/gdp_ARIMAgdp_currentUSD04.csv", index=True)

print("\n\n")





